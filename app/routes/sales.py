# app/sales.py
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, current_app, send_file
from flask_login import login_required, current_user
from app.models import db, Sale, SaleItem, Product, User, SystemSetting, PaymentMethod, SaleStatus, Expense, ExpenseCategory, ExpenseStatus, Payment, Customer, InvoiceStatus, ReturnTransaction, ReturnStatus
from sqlalchemy import func, and_
from datetime import datetime, timedelta
from io import BytesIO
import pdfkit
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import barcode
from barcode.writer import ImageWriter
import base64

sales_bp = Blueprint('sales', __name__, url_prefix='/sales')

# --------------------
# Utilities
# --------------------
from app.utils import format_currency

# --------------------
# Sales list
# --------------------
@sales_bp.route('/')
@login_required
def index():
    """
    List sales with optional filters:
      - start_date (YYYY-MM-DD)
      - end_date (YYYY-MM-DD)
      - user_id
      - page
    """
    page = request.args.get('page', 1, type=int)
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    user_id = request.args.get('user_id', type=int)

    query = Sale.query

    # parse dates (inclusive)
    start_date = None
    end_date = None
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            query = query.filter(Sale.created_at >= start_date)
        except ValueError:
            flash("Invalid start date format", "danger")

    if end_date_str:
        try:
            # use end of day for inclusive filter if you store times
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(Sale.created_at < end_date)
        except ValueError:
            flash("Invalid end date format", "danger")

    if user_id:
        query = query.filter(Sale.user_id == user_id)

    sales = query.order_by(Sale.created_at.desc()).paginate(page=page, per_page=10, error_out=False)
    users = User.query.order_by(User.username).all()

    return render_template(
        'sales/index.html',
        sales=sales,
        users=users,
        start_date=(start_date.date() if start_date else None),
        end_date=((end_date - timedelta(days=1)).date() if end_date else None),
        user_id=user_id,
        format_currency=format_currency,
        PaymentMethod=PaymentMethod,
        SaleStatus=SaleStatus
    )

# --------------------
# Sale detail
# --------------------
@sales_bp.route('/<int:sale_id>')
@login_required
def detail(sale_id):
    sale = Sale.query.get_or_404(sale_id)
    items = SaleItem.query.filter_by(sale_id=sale.id).order_by(SaleItem.id).all()

    # Safely load returns and void transaction (tables may not exist if migration pending)
    sale_returns = []
    void_txn = None
    try:
        sale_returns = sale.returns or []
    except Exception:
        pass
    try:
        void_txn = sale.void_transaction
    except Exception:
        pass

    return render_template(
        'sales/detail.html',
        sale=sale,
        items=items,
        sale_returns=sale_returns,
        void_txn=void_txn,
        format_currency=format_currency,
        PaymentMethod=PaymentMethod,
        SaleStatus=SaleStatus
    )

# --------------------
# Receipt (HTML view)
# --------------------
@sales_bp.route('/<int:sale_id>/receipt')
@login_required
def receipt(sale_id):
    sale = Sale.query.get_or_404(sale_id)
    items = SaleItem.query.filter_by(sale_id=sale.id).order_by(SaleItem.id).all()

    receipt_header = SystemSetting.get('receipt_header', 'Electronics Store POS System')
    receipt_footer = SystemSetting.get('receipt_footer', 'Thank you for your business!')

    # Generate Barcode
    barcode_b64 = None
    if sale.invoice_no:
        try:
            Code128 = barcode.get_barcode_class('code128')
            my_barcode = Code128(sale.invoice_no, writer=ImageWriter())
            buffer = BytesIO()
            my_barcode.write(buffer)
            barcode_b64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
        except Exception as e:
            current_app.logger.error(f"Barcode generation failed: {e}")

    return render_template(
        'sales/receipt.html',
        sale=sale,
        items=items,
        receipt_header=receipt_header,
        receipt_footer=receipt_footer,
        format_currency=format_currency,
        PaymentMethod=PaymentMethod,
        barcode_b64=barcode_b64
    )

# --------------------
# Receipt PDF (download)
# --------------------
@sales_bp.route('/<int:sale_id>/receipt/pdf')
@login_required
def receipt_pdf(sale_id):
    sale = Sale.query.get_or_404(sale_id)
    items = SaleItem.query.filter_by(sale_id=sale.id).order_by(SaleItem.id).all()

    receipt_header = SystemSetting.get('receipt_header', 'Electronics Store POS System')
    receipt_footer = SystemSetting.get('receipt_footer', 'Thank you for your business!')

    # Generate Barcode
    barcode_b64 = None
    if sale.invoice_no:
        try:
            Code128 = barcode.get_barcode_class('code128')
            my_barcode = Code128(sale.invoice_no, writer=ImageWriter())
            buffer = BytesIO()
            my_barcode.write(buffer)
            barcode_b64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
        except Exception as e:
            current_app.logger.error(f"Barcode generation failed: {e}")

    # Render PDF-friendly HTML template
    # Create a dedicated template 'sales/receipt_pdf.html' or reuse 'receipt.html' if that works well
    html = render_template(
        'sales/receipt_pdf.html',  # ensure this template exists and is printable
        sale=sale,
        items=items,
        receipt_header=receipt_header,
        receipt_footer=receipt_footer,
        format_currency=format_currency,
        PaymentMethod=PaymentMethod,
        barcode_b64=barcode_b64
    )

    # Generate PDF bytes
    try:
        path_wkhtmltopdf = current_app.config.get('WKHTMLTOPDF_PATH')
        config = pdfkit.configuration(wkhtmltopdf=path_wkhtmltopdf)
        
        options = {
            'page-size': 'A4',
            'margin-top': '10mm',
            'margin-right': '10mm',
            'margin-bottom': '10mm',
            'margin-left': '10mm',
            'encoding': "UTF-8"
        }
        
        pdf_bytes = pdfkit.from_string(html, False, configuration=config, options=options)
    except Exception as e:
        current_app.logger.exception("pdfkit failed to generate PDF")
        flash("PDF generation failed: " + str(e), "danger")
        return redirect(url_for('sales.detail', sale_id=sale.id))

    pdf_io = BytesIO(pdf_bytes)
    pdf_io.seek(0)
    filename = f"receipt_{sale.id}.pdf"

    # send as attachment for download
    return send_file(
        pdf_io,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )


# --------------------
# Recent Sales PDF
# --------------------
@sales_bp.route('/recent/pdf')
@login_required
def recent_pdf():
    """
    Generate PDF for the current list of sales (respecting filters).
    """
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    user_id = request.args.get('user_id', type=int)

    query = Sale.query

    # parse dates (inclusive)
    start_date = None
    end_date = None
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            query = query.filter(Sale.created_at >= start_date)
        except ValueError:
            flash("Invalid start date format", "danger")

    if end_date_str:
        try:
            # use end of day for inclusive filter if you store times
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(Sale.created_at < end_date)
        except ValueError:
            flash("Invalid end date format", "danger")

    if user_id:
        query = query.filter(Sale.user_id == user_id)

    # Get all matching sales (no pagination for PDF)
    sales = query.order_by(Sale.created_at.desc()).limit(500).all() # Limit to avoid massive PDFs

    # Calculate totals for the PDF
    total_subtotal = sum(s.subtotal for s in sales)
    total_tax = sum(s.tax_amount for s in sales)
    total_discount = sum(s.discount for s in sales)
    total_grand = sum(s.grand_total for s in sales)

    
    # Resolve user name for title if filtered
    user_filter_name = None
    if user_id:
        u = User.query.get(user_id)
        if u:
            user_filter_name = u.username

    html = render_template(
        'sales/list_pdf.html',
        sales=sales,
        generated_at=datetime.now().strftime('%Y-%m-%d %H:%M'),
        start_date=start_date_str,
        end_date=end_date_str,
        user_filter=user_filter_name,
        total_subtotal=total_subtotal,
        total_tax=total_tax,
        total_discount=total_discount,
        total_grand=total_grand,
        format_currency=format_currency,
        PaymentMethod=PaymentMethod,
        SaleStatus=SaleStatus
    )

    try:
        # Orientation Landscape for wider tables
        path_wkhtmltopdf = current_app.config.get('WKHTMLTOPDF_PATH')
        config = pdfkit.configuration(wkhtmltopdf=path_wkhtmltopdf)
        pdf_bytes = pdfkit.from_string(html, False, options={'orientation': 'Landscape'}, configuration=config)
    except Exception as e:
        current_app.logger.exception("pdfkit failed to generate sales list PDF")
        flash("PDF generation failed: " + str(e), "danger")
        return redirect(url_for('sales.index'))

    pdf_io = BytesIO(pdf_bytes)
    pdf_io.seek(0)
    filename = f"sales_list_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"

    return send_file(
        pdf_io,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )


# --------------------
# Sales List Excel Export
# --------------------
@sales_bp.route('/export/excel')
@login_required
def export_excel():
    """
    Generate Excel export for the current list of sales (respecting filters).
    """
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    user_id = request.args.get('user_id', type=int)

    query = Sale.query

    # parse dates (inclusive)
    start_date = None
    end_date = None
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            query = query.filter(Sale.created_at >= start_date)
        except ValueError:
            pass

    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(Sale.created_at < end_date)
        except ValueError:
            pass

    if user_id:
        query = query.filter(Sale.user_id == user_id)

    # Get all matching sales (limit to avoid massive exports)
    sales = query.order_by(Sale.created_at.desc()).limit(500).all()

    # Create Excel Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"

    # Headers
    headers = ['ID', 'Date', 'Cashier', 'Items', 'Subtotal', 'Tax', 'Discount', 'Total', 'Payment', 'Status']
    ws.append(headers)

    # Style Header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Add Data
    for sale in sales:
        try:
            payment = sale.payment_method.value
        except Exception:
            payment = str(sale.payment_method)
            
        # Fallback for empty payment method
        if not payment and sale.payments:
            try:
                payment = sale.payments[0].payment_method.value
            except:
                payment = str(sale.payments[0].payment_method)
        
        # User Request: "Store Credit" or empty should display as "Credit"
        if payment in ['Store Credit', 'STORE_CREDIT'] or not payment:
            payment = "Credit"
        
        try:
            status = sale.sale_status.value
        except Exception:
            status = str(sale.sale_status)

        row = [
            sale.id,
            sale.created_at.strftime('%Y-%m-%d %H:%M'),
            sale.user.username if sale.user else '-',
            len(sale.sale_items),
            float(sale.subtotal or 0),
            float(sale.tax_amount or 0),
            float(sale.discount or 0),
            float(sale.grand_total or 0),
            payment,
            status
        ]
        ws.append(row)

    # Auto-width columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save to BytesIO
    excel_io = BytesIO()
    wb.save(excel_io)
    excel_io.seek(0)
    
    filename = f"sales_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    return send_file(
        excel_io,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

# --------------------
# Reports page (HTML)
# --------------------
@sales_bp.route('/reports')
@login_required
def reports():
    if not current_user.has_role('Admin') and not current_user.has_role('Manager'):
        flash('Access denied', 'danger')
        return redirect(url_for('main.dashboard'))

    """
    Renders the reports page and injects arrays expected by the template:
      - daily_sales: list of {date: 'YYYY-MM-DD', sales: float} for each day in the range
      - payment_method_labels: [...], payment_method_values: [...]
      - top_products, user_sales, totals...
    """
    report_type = request.args.get('type', 'daily')
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')

    today = datetime.utcnow().date()
    start_date = None
    end_date = None

    # default ranges
    if report_type == 'daily':
        start_date = end_date = today
    elif report_type == 'weekly':
        end_date = today
        start_date = today - timedelta(days=7)
    elif report_type == 'monthly':
        end_date = today
        start_date = today.replace(day=1)

    # override with provided dates (custom)
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        except ValueError:
            flash('Invalid start date', 'danger')
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            flash('Invalid end date', 'danger')

    # fallback if dates missing
    if not start_date or not end_date:
        end_date = today
        start_date = today - timedelta(days=30)

    # Build daily sales query and fill gaps (EXCLUDE VOIDED)
    rows = db.session.query(
        func.date(Sale.created_at).label('date'),
        func.coalesce(func.sum(Sale.grand_total), 0).label('sales')
    ).filter(
        func.date(Sale.created_at) >= start_date,
        func.date(Sale.created_at) <= end_date,
        Sale.sale_status != SaleStatus.VOIDED
    ).group_by(func.date(Sale.created_at)).order_by(func.date(Sale.created_at)).all()

    row_map = {str(r.date): float(r.sales or 0) for r in rows}
    daily_sales = []
    cursor = start_date
    while cursor <= end_date:
        key = cursor.strftime('%Y-%m-%d')
        daily_sales.append({'date': key, 'sales': float(row_map.get(key, 0.0))})
        cursor += timedelta(days=1)

    # Payment methods (labels + values)
    # SUM ALL PAYMENTS in period (Cash Basis)
    pm_stats = db.session.query(
        Payment.payment_method,
        func.coalesce(func.sum(Payment.amount), 0).label('total')
    ).filter(
        func.date(Payment.created_at) >= start_date,
        func.date(Payment.created_at) <= end_date
    ).group_by(Payment.payment_method).all()

    pm_labels = []
    pm_values = []
    for r in pm_stats:
        try:
            label = r.payment_method.value
        except:
            label = str(r.payment_method)
        
        if label in ['Store Credit', 'STORE_CREDIT'] or not label:
            label = "Credit"
            
        pm_labels.append(label)
        pm_values.append(float(r.total or 0))

    # Totals and aggregates (EXCLUDE VOIDED)
    # Invoiced Total (Accrual Basis - what we sold)
    total_sales_count = db.session.query(func.count(Sale.id)).filter(
        func.date(Sale.created_at) >= start_date,
        func.date(Sale.created_at) <= end_date,
        Sale.sale_status != SaleStatus.VOIDED
    ).scalar() or 0

    total_invoiced = db.session.query(func.coalesce(func.sum(Sale.grand_total), 0)).filter(
        func.date(Sale.created_at) >= start_date,
        func.date(Sale.created_at) <= end_date,
        Sale.sale_status != SaleStatus.VOIDED
    ).scalar() or 0

    total_items = db.session.query(func.coalesce(func.sum(SaleItem.quantity_sold), 0)).join(Sale, SaleItem.sale).filter(
        func.date(Sale.created_at) >= start_date,
        func.date(Sale.created_at) <= end_date,
        Sale.sale_status != SaleStatus.VOIDED
    ).scalar() or 0

    low_stock_count = Product.query.filter(Product.quantity_in_stock <= Product.low_stock_threshold).count()

    # --- CASH FLOW ANALYSIS ---
    
    # 1. Total Collected (Cash Basis - what we received)
    total_collected = db.session.query(func.coalesce(func.sum(Payment.amount), 0)).filter(
        func.date(Payment.created_at) >= start_date,
        func.date(Payment.created_at) <= end_date
    ).scalar() or 0

    # 2. COGS (Cost of Goods Sold) - EXCLUDE VOIDED
    cogs = db.session.query(
        func.coalesce(func.sum(SaleItem.quantity_sold * Product.cost_price), 0)
    ).join(Product).join(Sale, SaleItem.sale).filter(
        func.date(Sale.created_at) >= start_date,
        func.date(Sale.created_at) <= end_date,
        Sale.sale_status != SaleStatus.VOIDED
    ).scalar() or 0

    # 3. Real Profit - EXCLUDE VOIDED
    period_sales = Sale.query.filter(
        func.date(Sale.created_at) >= start_date,
        func.date(Sale.created_at) <= end_date,
        Sale.sale_status != SaleStatus.VOIDED
    ).all()
    gross_profit = sum(sale.real_profit for sale in period_sales)

    # --- VOIDED SALES METRICS ---
    voided_count = db.session.query(func.count(Sale.id)).filter(
        func.date(Sale.created_at) >= start_date,
        func.date(Sale.created_at) <= end_date,
        Sale.sale_status == SaleStatus.VOIDED
    ).scalar() or 0

    voided_value = db.session.query(func.coalesce(func.sum(Sale.grand_total), 0)).filter(
        func.date(Sale.created_at) >= start_date,
        func.date(Sale.created_at) <= end_date,
        Sale.sale_status == SaleStatus.VOIDED
    ).scalar() or 0

    # --- RETURNS METRICS ---
    return_count = db.session.query(func.count(ReturnTransaction.id)).filter(
        func.date(ReturnTransaction.created_at) >= start_date,
        func.date(ReturnTransaction.created_at) <= end_date,
        ReturnTransaction.status == ReturnStatus.COMPLETED
    ).scalar() or 0

    return_value = db.session.query(
        func.coalesce(func.sum(ReturnTransaction.refund_total), 0)
    ).filter(
        func.date(ReturnTransaction.created_at) >= start_date,
        func.date(ReturnTransaction.created_at) <= end_date,
        ReturnTransaction.status == ReturnStatus.COMPLETED
    ).scalar() or 0

    # Net revenue = invoiced - returns
    net_revenue = float(total_invoiced) - float(return_value)

    # Average Order Value (based on Invoiced Total)
    avg_order_value = float(total_invoiced) / float(total_sales_count) if total_sales_count > 0 else 0

    # Top products (EXCLUDE VOIDED)
    top_products = db.session.query(
        Product.name,
        func.coalesce(func.sum(SaleItem.quantity_sold), 0).label('total_sold'),
        func.coalesce(func.sum(SaleItem.total_price), 0).label('total_revenue')
    ).join(SaleItem).join(Sale).filter(
        func.date(Sale.created_at) >= start_date,
        func.date(Sale.created_at) <= end_date,
        Sale.sale_status != SaleStatus.VOIDED
    ).group_by(Product.id).order_by(func.sum(SaleItem.quantity_sold).desc()).limit(10).all()

    # Sales by cashier
    user_sales = db.session.query(
        User.username,
        func.count(Sale.id).label('count'),
        func.coalesce(func.sum(Sale.grand_total), 0).label('total')
    ).join(Sale).filter(
        func.date(Sale.created_at) >= start_date,
        func.date(Sale.created_at) <= end_date
    ).group_by(User.id).order_by(func.sum(Sale.grand_total).desc()).all()

    # --- EXPENSES INTEGRATION ---
    paid_expenses = db.session.query(func.coalesce(func.sum(Expense.amount), 0)).filter(
        Expense.date >= start_date,
        Expense.date <= end_date,
        Expense.status == ExpenseStatus.PAID
    ).scalar() or 0

    pending_expenses = db.session.query(func.coalesce(func.sum(Expense.amount), 0)).filter(
        Expense.date >= start_date,
        Expense.date <= end_date,
        Expense.status == ExpenseStatus.PENDING
    ).scalar() or 0

    total_expenses = float(paid_expenses) + float(pending_expenses)

    # Net Income = Net Revenue - Total Expenses
    net_profit = net_revenue - float(total_expenses)

    # Expense Distribution by Category (PAID only for chart accuracy)
    exp_stats = db.session.query(
        ExpenseCategory.name,
        func.coalesce(func.sum(Expense.amount), 0).label('total'),
        ExpenseCategory.color
    ).join(Expense).filter(
        Expense.date >= start_date,
        Expense.date <= end_date,
        Expense.status == ExpenseStatus.PAID
    ).group_by(ExpenseCategory.id).all()

    exp_labels = [r[0] for r in exp_stats]
    exp_values = [float(r[1]) for r in exp_stats]
    exp_colors = [r[2] for r in exp_stats]

    # Recent Expenses
    recent_expenses = Expense.query.filter(
        Expense.date >= start_date,
        Expense.date <= end_date
    ).order_by(Expense.date.desc()).limit(10).all()


    # Sales by cashier
    user_sales = db.session.query(
        User.username,
        func.count(Sale.id).label('count'),
        func.coalesce(func.sum(Sale.grand_total), 0).label('total')
    ).join(Sale).filter(
        func.date(Sale.created_at) >= start_date,
        func.date(Sale.created_at) <= end_date
    ).group_by(User.id).order_by(func.sum(Sale.grand_total).desc()).all()

    # Top Customers (Added for Leaderboard)
    top_customers = db.session.query(
        Customer.full_name,
        func.count(Sale.id).label('tx_count'),
        func.sum(Sale.grand_total).label('total_spent')
    ).join(Sale).filter(
        func.date(Sale.created_at) >= start_date,
        func.date(Sale.created_at) <= end_date,
        Sale.customer_id.isnot(None) 
    ).group_by(Customer.id).order_by(func.sum(Sale.grand_total).desc()).limit(10).all()

    # --- INVENTORY VALUATION (ASSETS) ---
    stock_stats = db.session.query(
        func.coalesce(func.sum(Product.quantity_in_stock), 0).label('total_qty'),
        func.coalesce(func.sum(Product.quantity_in_stock * Product.cost_price), 0).label('total_cost'),
        func.coalesce(func.sum(Product.quantity_in_stock * Product.selling_price), 0).label('total_retail')
    ).filter(Product.quantity_in_stock > 0).first()

    if stock_stats:
        inventory_qty = stock_stats.total_qty or 0
        inventory_cost = stock_stats.total_cost or 0
        inventory_retail = stock_stats.total_retail or 0
    else:
        inventory_qty = 0
        inventory_cost = 0
        inventory_retail = 0

    inventory_profit = float(inventory_retail) - float(inventory_cost)

    return render_template(
        'sales/reports.html',
        inventory_qty=inventory_qty,
        inventory_cost=inventory_cost,
        inventory_retail=inventory_retail,
        inventory_profit=inventory_profit,
        report_type=report_type,
        start_date=start_date,
        end_date=end_date,
        total_sales=int(total_sales_count),
        total_invoiced=float(total_invoiced),
        total_collected=float(total_collected),
        total_items=int(total_items),
        cogs=float(cogs),
        gross_profit=gross_profit,  # This is now REAL profit
        avg_order_value=avg_order_value,
        # ... rest same ...
        paid_expenses=float(paid_expenses),
        pending_expenses=float(pending_expenses),
        total_expenses=float(total_expenses),
        net_profit=net_profit,
        low_stock_count=int(low_stock_count),
        daily_sales=daily_sales,
        payment_method_labels=pm_labels,
        payment_method_values=pm_values,
        expense_labels=exp_labels,
        expense_values=exp_values,
        expense_colors=exp_colors,
        recent_expenses=recent_expenses,
        top_products=top_products,
        user_sales=user_sales,
        top_customers=top_customers,
        voided_count=voided_count,
        voided_value=float(voided_value),
        return_count=return_count,
        return_value=float(return_value),
        net_revenue=net_revenue,
        format_currency=format_currency
    )

# --------------------
# Reports JSON (AJAX)
# --------------------
@sales_bp.route('/reports/data')
@login_required
def reports_data():
    if not current_user.has_role('Admin') and not current_user.has_role('Manager'):
        return jsonify({'error': 'Access denied'}), 403

    """
    Minimal JSON endpoint used by older versions / AJAX.
    Returns:
      { daily_sales: [{date, sales}, ...], payment_methods: { labels:[], values:[] } }
    Default: last 30 days
    """
    today = datetime.utcnow().date()
    start = request.args.get('start_date')
    end = request.args.get('end_date')
    report_type = request.args.get('type', 'daily')

    if start and end:
        try:
            start_date = datetime.strptime(start, '%Y-%m-%d').date()
            end_date = datetime.strptime(end, '%Y-%m-%d').date()
        except ValueError:
            start_date = today - timedelta(days=30)
            end_date = today
    else:
        # fallback based on type
        if report_type == 'weekly':
            start_date = today - timedelta(days=7)
            end_date = today
        elif report_type == 'monthly':
            start_date = today - timedelta(days=30)
            end_date = today
        else:
            start_date = today - timedelta(days=30)
            end_date = today

    # daily sales
    rows = db.session.query(
        func.date(Sale.created_at).label('date'),
        func.coalesce(func.sum(Sale.grand_total), 0).label('sales')
    ).filter(func.date(Sale.created_at) >= start_date).filter(func.date(Sale.created_at) <= end_date).group_by(func.date(Sale.created_at)).order_by(func.date(Sale.created_at)).all()

    daily = [{'date': str(r.date), 'sales': float(r.sales or 0)} for r in rows]

    pm_rows = db.session.query(
        Sale.payment_method,
        func.coalesce(func.sum(Sale.grand_total), 0).label('total')
    ).filter(func.date(Sale.created_at) >= start_date).filter(func.date(Sale.created_at) <= end_date).group_by(Sale.payment_method).all()

    pm_labels = []
    pm_values = []
    for r in pm_rows:
        try:
            label = r.payment_method.value
        except Exception:
            label = str(r.payment_method)
        pm_labels.append(label)
        pm_values.append(float(r.total or 0))

    return jsonify({
        'daily_sales': daily,
        'payment_methods': {'labels': pm_labels, 'values': pm_values}
    })


# --------------------
# Reports PDF
# --------------------
@sales_bp.route('/reports/pdf')
@login_required
def reports_pdf():
    if not current_user.has_role('Admin') and not current_user.has_role('Manager'):
        flash('Access denied', 'danger')
        return redirect(url_for('main.dashboard'))

    """
    Generate PDF for the reports page.
    """
    report_type = request.args.get('type', 'daily')
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')

    today = datetime.utcnow().date()
    start_date = None
    end_date = None

    # default ranges
    if report_type == 'daily':
        start_date = end_date = today
    elif report_type == 'weekly':
        end_date = today
        start_date = today - timedelta(days=7)
    elif report_type == 'monthly':
        end_date = today
        start_date = today.replace(day=1)

    # override with provided dates (custom)
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass

    # fallback
    if not start_date or not end_date:
        end_date = today
        start_date = today - timedelta(days=30)
        
    # Re-run queries for the report data
    # 1. Daily Sales
    rows = db.session.query(
        func.date(Sale.created_at).label('date'),
        func.coalesce(func.sum(Sale.grand_total), 0).label('sales')
    ).filter(
        func.date(Sale.created_at) >= start_date,
        func.date(Sale.created_at) <= end_date
    ).group_by(func.date(Sale.created_at)).order_by(func.date(Sale.created_at)).all()
    
    daily_sales = [{'date': str(r.date), 'sales': float(r.sales or 0)} for r in rows]

    # 2. Aggregates
    total_sales = db.session.query(func.count(Sale.id)).filter(
        func.date(Sale.created_at) >= start_date,
        func.date(Sale.created_at) <= end_date
    ).scalar() or 0

    total_revenue = db.session.query(func.coalesce(func.sum(Sale.grand_total), 0)).filter(
        func.date(Sale.created_at) >= start_date,
        func.date(Sale.created_at) <= end_date
    ).scalar() or 0

    total_items = db.session.query(func.coalesce(func.sum(SaleItem.quantity_sold), 0)).join(Sale, SaleItem.sale).filter(
        func.date(Sale.created_at) >= start_date,
        func.date(Sale.created_at) <= end_date
    ).scalar() or 0

    # 3. Top Products
    top_products = db.session.query(
        Product.name,
        func.coalesce(func.sum(SaleItem.quantity_sold), 0).label('total_sold'),
        func.coalesce(func.sum(SaleItem.total_price), 0).label('total_revenue')
    ).join(SaleItem).join(Sale).filter(
        func.date(Sale.created_at) >= start_date,
        func.date(Sale.created_at) <= end_date
    ).group_by(Product.id).order_by(func.sum(SaleItem.quantity_sold).desc()).limit(10).all()

    # 4. Sales by cashier
    user_sales = db.session.query(
        User.username,
        func.count(Sale.id).label('count'),
        func.coalesce(func.sum(Sale.grand_total), 0).label('total')
    ).join(Sale).filter(
        func.date(Sale.created_at) >= start_date,
        func.date(Sale.created_at) <= end_date
    ).group_by(User.id).order_by(func.sum(Sale.grand_total).desc()).all()

    # 5. COGS for PDF
    # 5. Realized Profit (Strict Cash Basis)
    period_sales = Sale.query.filter(
        func.date(Sale.created_at) >= start_date,
        func.date(Sale.created_at) <= end_date
    ).all()

    realized_revenue = 0
    realized_cogs = 0
    for sale in period_sales:
        if sale.invoice_status == InvoiceStatus.PAID:
            sale_revenue = sale.grand_total 
            sale_cogs = sum(item.quantity_sold * item.product.cost_price for item in sale.sale_items)
            realized_revenue += float(sale_revenue)
            realized_cogs += float(sale_cogs)

    realized_gross_profit = realized_revenue - realized_cogs
    
    # 6. Expenses
    paid_expenses = db.session.query(func.coalesce(func.sum(Expense.amount), 0)).filter(
        Expense.date >= start_date,
        Expense.date <= end_date,
        Expense.status == ExpenseStatus.PAID
    ).scalar() or 0

    pending_expenses = db.session.query(func.coalesce(func.sum(Expense.amount), 0)).filter(
        Expense.date >= start_date,
        Expense.date <= end_date,
        Expense.status == ExpenseStatus.PENDING
    ).scalar() or 0

    total_expenses = float(paid_expenses) + float(pending_expenses)
    net_profit = float(realized_gross_profit) - float(paid_expenses)

    # 7. Top Customers
    top_customers = db.session.query(
        Customer.full_name,
        func.count(Sale.id).label('tx_count'),
        func.sum(Sale.grand_total).label('total_spent')
    ).join(Sale).filter(
        func.date(Sale.created_at) >= start_date,
        func.date(Sale.created_at) <= end_date,
        Sale.customer_id.isnot(None) 
    ).group_by(Customer.id).order_by(func.sum(Sale.grand_total).desc()).limit(10).all()

    # 8. Avg Order Value implies total invoiced / count
    avg_order_value = float(total_revenue) / float(total_sales) if total_sales > 0 else 0

    # Get Company Info
    company_name = SystemSetting.get('company_name', 'Electronics POS')
    company_logo = SystemSetting.get('company_logo')

    period_expenses = Expense.query.filter(
        Expense.date >= start_date,
        Expense.date <= end_date
    ).order_by(Expense.date.desc()).all()

    # Inventory Valuation
    stock_stats = db.session.query(
        func.coalesce(func.sum(Product.quantity_in_stock), 0).label('total_qty'),
        func.coalesce(func.sum(Product.quantity_in_stock * Product.cost_price), 0).label('total_cost'),
        func.coalesce(func.sum(Product.quantity_in_stock * Product.selling_price), 0).label('total_retail')
    ).filter(Product.quantity_in_stock > 0).first()

    if stock_stats:
        inventory_qty = stock_stats.total_qty or 0
        inventory_cost = stock_stats.total_cost or 0
        inventory_retail = stock_stats.total_retail or 0
    else:
        inventory_qty = 0
        inventory_cost = 0
        inventory_retail = 0

    inventory_profit = float(inventory_retail) - float(inventory_cost)

    html = render_template(
        'sales/reports_pdf.html',
        generated_at=datetime.now().strftime('%d %b %Y, %I:%M %p'),
        start_date=start_date,
        end_date=end_date,
        report_type=report_type,
        company_name=company_name,
        company_logo=company_logo,
        
        total_sales=total_sales,
        total_revenue=total_revenue,
        cogs=realized_cogs,  # Pass realized cogs
        gross_profit=realized_gross_profit, # Pass realized profit
        avg_order_value=avg_order_value,
        
        paid_expenses=float(paid_expenses),
        pending_expenses=float(pending_expenses),
        total_expenses=total_expenses,
        net_profit=net_profit,
        
        total_items=total_items,
        # daily_sales=daily_sales, # Not used in PDF usually?
        top_products=top_products,
        user_sales=user_sales,
        period_expenses=period_expenses,
        top_customers=top_customers,
        inventory_qty=inventory_qty,
        inventory_cost=inventory_cost,
        inventory_retail=inventory_retail,
        inventory_profit=inventory_profit,
        format_currency=format_currency
    )

    try:
        path_wkhtmltopdf = current_app.config.get('WKHTMLTOPDF_PATH')
        config = pdfkit.configuration(wkhtmltopdf=path_wkhtmltopdf)
        pdf_bytes = pdfkit.from_string(html, False, configuration=config)
    except Exception as e:
        current_app.logger.exception("pdfkit failed to generate reports PDF")
        flash("PDF generation failed: " + str(e), "danger")
        return redirect(url_for('sales.reports'))

    pdf_io = BytesIO(pdf_bytes)
    pdf_io.seek(0)
    filename = f"sales_report_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"

    return send_file(
        pdf_io,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )


# --------------------
# Reports Excel Export
# --------------------
@sales_bp.route('/reports/excel')
@login_required
def reports_excel():
    if not current_user.has_role('Admin') and not current_user.has_role('Manager'):
        flash('Access denied', 'danger')
        return redirect(url_for('main.dashboard'))

    """
    Generate Excel export for the reports page with multiple sheets.
    """
    report_type = request.args.get('type', 'daily')
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')

    today = datetime.utcnow().date()
    start_date = None
    end_date = None

    # default ranges
    if report_type == 'daily':
        start_date = end_date = today
    elif report_type == 'weekly':
        end_date = today
        start_date = today - timedelta(days=7)
    elif report_type == 'monthly':
        end_date = today
        start_date = today.replace(day=1)

    # override with provided dates (custom)
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass

    # fallback
    if not start_date or not end_date:
        end_date = today
        start_date = today - timedelta(days=30)

    # Gather report data
    # 1. Summary stats
    total_sales = db.session.query(func.count(Sale.id)).filter(
        func.date(Sale.created_at) >= start_date,
        func.date(Sale.created_at) <= end_date
    ).scalar() or 0

    total_revenue = db.session.query(func.coalesce(func.sum(Sale.grand_total), 0)).filter(
        func.date(Sale.created_at) >= start_date,
        func.date(Sale.created_at) <= end_date
    ).scalar() or 0

    total_items = db.session.query(func.coalesce(func.sum(SaleItem.quantity_sold), 0)).join(Sale, SaleItem.sale).filter(
        func.date(Sale.created_at) >= start_date,
        func.date(Sale.created_at) <= end_date
    ).scalar() or 0

    # 3. Realized Profit (Strict Cash Basis: Only Fully Paid Invoices)
    # Fetch all sales in period first
    period_sales = Sale.query.filter(
        func.date(Sale.created_at) >= start_date,
        func.date(Sale.created_at) <= end_date
    ).all()

    # Calculate Realized Revenue & COGS from PAID sales only
    realized_revenue = 0
    realized_cogs = 0
    
    # We can also track "Potential Profit" from unpaid/partial for insights if needed, 
    # but user requested STRICT realized profit.
    for sale in period_sales:
        # Check for strict PAID status
        if sale.invoice_status == InvoiceStatus.PAID:
            sale_revenue = sale.grand_total # Or subtotal? Using grand_total (Revenue)
            # Calculate sale COGS
            sale_cogs = sum(item.quantity_sold * item.product.cost_price for item in sale.sale_items)
            
            realized_revenue += float(sale_revenue)
            realized_cogs += float(sale_cogs)

    # Realized Gross Profit
    realized_gross_profit = realized_revenue - realized_cogs

    # Average Order Value
    avg_order_value = float(total_revenue) / float(total_sales) if total_sales > 0 else 0

    # Expenses
    paid_expenses = db.session.query(func.coalesce(func.sum(Expense.amount), 0)).filter(
        Expense.date >= start_date,
        Expense.date <= end_date,
        Expense.status == ExpenseStatus.PAID
    ).scalar() or 0

    pending_expenses = db.session.query(func.coalesce(func.sum(Expense.amount), 0)).filter(
        Expense.date >= start_date,
        Expense.date <= end_date,
        Expense.status == ExpenseStatus.PENDING
    ).scalar() or 0

    total_expenses = float(paid_expenses) + float(pending_expenses)
    net_profit = float(realized_gross_profit) - float(paid_expenses)

    # Daily Sales
    daily_rows = db.session.query(
        func.date(Sale.created_at).label('date'),
        func.coalesce(func.sum(Sale.grand_total), 0).label('sales')
    ).filter(
        func.date(Sale.created_at) >= start_date,
        func.date(Sale.created_at) <= end_date,
        Sale.sale_status != SaleStatus.VOIDED
    ).group_by(func.date(Sale.created_at)).order_by(func.date(Sale.created_at)).all()

    daily_row_map = {str(r.date): float(r.sales or 0) for r in daily_rows}
    daily_sales = []
    cursor = start_date
    while cursor <= end_date:
        key = cursor.strftime('%Y-%m-%d')
        daily_sales.append({'date': key, 'sales': float(daily_row_map.get(key, 0.0))})
        cursor += timedelta(days=1)

    # Top products
    top_products = db.session.query(
        Product.name,
        func.coalesce(func.sum(SaleItem.quantity_sold), 0).label('total_sold'),
        func.coalesce(func.sum(SaleItem.total_price), 0).label('total_revenue')
    ).join(SaleItem).join(Sale).filter(
        func.date(Sale.created_at) >= start_date,
        func.date(Sale.created_at) <= end_date
    ).group_by(Product.id).order_by(func.sum(SaleItem.quantity_sold).desc()).limit(10).all()

    # Top Customers
    top_customers = db.session.query(
        Customer.full_name,
        func.count(Sale.id).label('tx_count'),
        func.sum(Sale.grand_total).label('total_spent')
    ).join(Sale).filter(
        func.date(Sale.created_at) >= start_date,
        func.date(Sale.created_at) <= end_date,
        Sale.customer_id.isnot(None) 
    ).group_by(Customer.id).order_by(func.sum(Sale.grand_total).desc()).limit(10).all()

    # Sales by cashier
    user_sales = db.session.query(
        User.username,
        func.count(Sale.id).label('count'),
        func.coalesce(func.sum(Sale.grand_total), 0).label('total')
    ).join(Sale).filter(
        func.date(Sale.created_at) >= start_date,
        func.date(Sale.created_at) <= end_date
    ).group_by(User.id).order_by(func.sum(Sale.grand_total).desc()).all()

    # Create Excel Workbook
    wb = openpyxl.Workbook()
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    money_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    ws_summary.append(["Sales Report Summary"])
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary.append([f"Period: {start_date} to {end_date}"])
    ws_summary.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    ws_summary.append([])
    
    summary_data = [
        ["Metric", "Value"],
        ["Total Transactions", int(total_sales)],
        ["Total Revenue", float(total_revenue)],
        ["Total Items Sold", int(total_items)],
        ["Cost of Goods Sold (COGS)", float(realized_cogs)],
        ["Gross Profit", float(realized_gross_profit)],
        ["Average Order Value", float(avg_order_value)],
        ["Paid Expenses", float(paid_expenses)],
        ["Pending Expenses", float(pending_expenses)],
        ["Net Profit", float(net_profit)]
    ]
    
    for row in summary_data:
        ws_summary.append(row)
    
    for cell in ws_summary[5]:
        cell.font = header_font
        cell.fill = header_fill

    # Sheet 2: Daily Sales
    ws_daily = wb.create_sheet("Daily Sales")
    ws_daily.append(["Date", "Revenue"])
    for cell in ws_daily[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    for d in daily_sales:
        ws_daily.append([d['date'], d['sales']])

    # Sheet 3: Top Products
    ws_products = wb.create_sheet("Top Products")
    ws_products.append(["Product Name", "Quantity Sold", "Revenue"])
    for cell in ws_products[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    for p in top_products:
        ws_products.append([p[0], int(p[1]), float(p[2])])

    # Sheet 4: Cashier Performance
    ws_cashiers = wb.create_sheet("Cashier Performance")
    ws_cashiers.append(["Cashier", "Transactions", "Total Revenue"])
    for cell in ws_cashiers[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    for u in user_sales:
        ws_cashiers.append([u[0], int(u[1]), float(u[2])])

    # Auto-width for all sheets
    for ws in wb.worksheets:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

    # Save to BytesIO
    excel_io = BytesIO()
    wb.save(excel_io)
    excel_io.seek(0)
    
    filename = f"sales_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    return send_file(
        excel_io,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
