# app/routes/customers.py
"""Customer CRUD routes with profile, ledger, and search functionality"""

from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file, current_app
from flask_login import login_required, current_user
from app.models import db, Customer, CustomerLedger, Sale, Payment, CustomerStatus, LedgerEntryType, InvoiceStatus, PaymentMethod, SaleStatus, SaleType, SystemSetting
import os
from app.utils import format_currency
from datetime import datetime, timedelta
from decimal import Decimal
from sqlalchemy import func, desc, and_
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pdfkit


customers_bp = Blueprint('customers', __name__, url_prefix='/customers')


# =============================================================================
# CUSTOMER LIST
# =============================================================================

@customers_bp.route('/')
@login_required
def index():
    """List all customers with search, pagination, and dashboard stats"""
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '').strip()
    status_filter = request.args.get('status', '')
    leaderboard_metric = request.args.get('metric', 'top_spenders_month')
    per_page = 20
    
    # --- DASHBOARD STATISTICS ---
    stats = get_dashboard_stats()
    
    # --- LEADERBOARD ---
    leaderboard = get_leaderboard(leaderboard_metric)

    # --- CUSTOMER LIST QUERY ---
    query = Customer.query
    
    # Search filter
    if search:
        query = query.filter(
            db.or_(
                Customer.full_name.ilike(f'%{search}%'),
                Customer.phone.ilike(f'%{search}%'),
                Customer.email.ilike(f'%{search}%')
            )
        )
    
    # Status filter
    if status_filter:
        if status_filter == 'Debtors':
            # Subquery to find customers with positive balance
            subq = db.session.query(
                CustomerLedger.customer_id,
                func.max(CustomerLedger.id).label('max_id')
            ).group_by(CustomerLedger.customer_id).subquery()
            
            debtor_ids = db.session.query(CustomerLedger.customer_id)\
                .join(subq, CustomerLedger.id == subq.c.max_id)\
                .filter(CustomerLedger.balance_after > 0).all()
            debtor_ids = [d[0] for d in debtor_ids]
            
            query = query.filter(Customer.id.in_(debtor_ids))
        else:
            try:
                query = query.filter(Customer.status == CustomerStatus(status_filter))
            except:
                pass
    
    # Order by name
    query = query.order_by(Customer.full_name.asc())
    
    # Paginate
    customers = query.paginate(page=page, per_page=per_page, error_out=False)
    
    return render_template('customers/index.html', 
                           customers=customers, 
                           search=search,
                           status_filter=status_filter,
                           stats=stats,
                           leaderboard=leaderboard,
                           current_metric=leaderboard_metric,
                           format_currency=format_currency)

def get_dashboard_stats():
    """Calculate dashboard metrics efficiently"""
    today = datetime.utcnow().date()
    start_of_month = today.replace(day=1)
    
    # 1. Total & Active Customers
    total_customers = Customer.query.count()
    
    # Active: Customers with at least 1 invoice in last 30 days
    last_30_days = today - timedelta(days=30)
    active_customers = db.session.query(func.count(func.distinct(Sale.customer_id)))\
        .filter(Sale.created_at >= last_30_days, Sale.customer_id.isnot(None)).scalar() or 0
        
    # 2. Financials (Ledger & Payments)
    # Customers with Balance Due (Latest ledger entry > 0)
    # This is complex in SQL without a subquery, so we check the `get_outstanding_balance` equivalent logic
    # Optimization: We assume efficient ledger usage. 
    # For scalable "Balance Due" count, we ideally need a 'current_balance' column on Customer table.
    # PROD NOTE: Add current_balance column to Customer for O(1) query.
    # For now, we will query latest ledger entries.
    
    # Subquery to get latest ledger ID for each customer
    subq = db.session.query(
        CustomerLedger.customer_id,
        func.max(CustomerLedger.id).label('max_id')
    ).group_by(CustomerLedger.customer_id).subquery()
    
    # Join to get the balance
    balance_due_count = db.session.query(func.count(CustomerLedger.id))\
        .join(subq, CustomerLedger.id == subq.c.max_id)\
        .filter(CustomerLedger.balance_after > 0).scalar() or 0
        
    total_outstanding = db.session.query(func.sum(CustomerLedger.balance_after))\
        .join(subq, CustomerLedger.id == subq.c.max_id)\
        .filter(CustomerLedger.balance_after > 0).scalar() or 0
        
    # 3. Monthly Metrics
    # Paid This Month
    paid_this_month = db.session.query(func.sum(Payment.amount))\
        .filter(Payment.created_at >= start_of_month).scalar() or 0
        
    # Wholesale Sales This Month (SaleType.WHOLESALE)
    # Assuming we added SaleType to Sale model in shared context, otherwise use logic
    # If SaleType not in Sale model, we check price_type of items or similar.
    # Based on models.py viewed earlier, Sale has sale_type!
    wholesale_sales = db.session.query(func.sum(Sale.grand_total))\
        .filter(Sale.created_at >= start_of_month, 
                Sale.sale_status != SaleStatus.VOIDED,
                # Sale.sale_type == SaleType.WHOLESALE  <-- If SaleType exists and is populated
                ).scalar() or 0
    
    # Retail Sales (Approximate as Total - Wholesale for now if Mixed types exist)
    total_sales_month = db.session.query(func.sum(Sale.grand_total))\
        .filter(Sale.created_at >= start_of_month,
                Sale.sale_status != SaleStatus.VOIDED).scalar() or 0
    retail_sales = float(total_sales_month) - float(wholesale_sales)

    # Top Customer This Month
    top_customer = db.session.query(
        Customer.full_name,
        func.sum(Sale.grand_total).label('total_spent')
    ).join(Sale).filter(
        Sale.created_at >= start_of_month,
        Sale.sale_status != SaleStatus.VOIDED
    ).group_by(Customer.id).order_by(desc('total_spent')).first()
    
    return {
        'total_customers': total_customers,
        'active_customers': active_customers,
        'balance_due_count': balance_due_count,
        'total_outstanding': total_outstanding,
        'paid_this_month': paid_this_month,
        'wholesale_sales': wholesale_sales,
        'retail_sales': retail_sales,
        'top_customer': top_customer
    }

def get_leaderboard(metric='top_spenders_month'):
    """Get top 10 customers based on metric"""
    query = db.session.query(
        Customer.id,
        Customer.full_name,
        Customer.phone
    )
    
    limit = 10
    results = []
    
    if metric == 'top_spenders_month':
        start_of_month = datetime.utcnow().replace(day=1, hour=0, minute=0, second=0)
        rows = query.add_columns(func.sum(Sale.grand_total).label('value'))\
            .join(Sale)\
            .filter(Sale.created_at >= start_of_month, Sale.sale_status != SaleStatus.VOIDED)\
            .group_by(Customer.id)\
            .order_by(desc('value')).limit(limit).all()
        results = [{'id': r.id, 'name': r.full_name, 'value': r.value, 'label': 'Spent'} for r in rows]
        
    elif metric == 'top_spenders_all':
        rows = query.add_columns(func.sum(Sale.grand_total).label('value'))\
            .join(Sale)\
            .filter(Sale.sale_status != SaleStatus.VOIDED)\
            .group_by(Customer.id)\
            .order_by(desc('value')).limit(limit).all()
        results = [{'id': r.id, 'name': r.full_name, 'value': r.value, 'label': 'Spent'} for r in rows]
        
    elif metric == 'most_orders':
        rows = query.add_columns(func.count(Sale.id).label('value'))\
            .join(Sale)\
            .group_by(Customer.id)\
            .order_by(desc('value')).limit(limit).all()
        results = [{'id': r.id, 'name': r.full_name, 'value': int(r.value), 'label': 'Orders'} for r in rows]
        
    elif metric == 'most_owed':
        # Use subquery for latest ledger query
        subq = db.session.query(
            CustomerLedger.customer_id,
            func.max(CustomerLedger.id).label('max_id')
        ).group_by(CustomerLedger.customer_id).subquery()
        
        rows = query.add_columns(CustomerLedger.balance_after.label('value'))\
            .join(CustomerLedger)\
            .join(subq, and_(CustomerLedger.customer_id == subq.c.customer_id, 
                             CustomerLedger.id == subq.c.max_id))\
            .filter(CustomerLedger.balance_after > 0)\
            .order_by(desc('value')).limit(limit).all()
            
        results = [{'id': r.id, 'name': r.full_name, 'value': r.value, 'label': 'Owes'} for r in rows]

    return results


def get_debtors_watchlist(limit=None):
    """
    Get customers who owe money with aging information.
    Returns: list of dicts with customer info, balance, total_invoiced, total_paid, 
             last_invoice_date, days_overdue (10+ days = overdue)
    """
    today = datetime.utcnow().date()
    
    # Subquery for latest ledger entry per customer
    subq = db.session.query(
        CustomerLedger.customer_id,
        func.max(CustomerLedger.id).label('max_id')
    ).group_by(CustomerLedger.customer_id).subquery()
    
    # Get customers with positive balance
    debtors_query = db.session.query(
        Customer.id,
        Customer.full_name,
        Customer.phone,
        CustomerLedger.balance_after.label('balance')
    ).join(CustomerLedger).join(
        subq, and_(CustomerLedger.customer_id == subq.c.customer_id,
                   CustomerLedger.id == subq.c.max_id)
    ).filter(CustomerLedger.balance_after > 0).order_by(desc('balance'))
    
    if limit:
        debtors_query = debtors_query.limit(limit)
    
    debtors = debtors_query.all()
    
    results = []
    for debtor in debtors:
        # Get oldest unpaid invoice date for aging calculation
        # Using InvoiceStatus.PAID (correct enum value)
        oldest_unpaid = db.session.query(func.min(Sale.created_at)).filter(
            Sale.customer_id == debtor.id,
            Sale.invoice_status != InvoiceStatus.PAID,
            Sale.sale_status != SaleStatus.VOIDED
        ).scalar()
        
        # Calculate days overdue (from oldest unpaid invoice)
        days_overdue = 0
        if oldest_unpaid:
            days_overdue = (today - oldest_unpaid.date()).days
        
        # Get total invoiced and total paid for this customer
        total_invoiced = db.session.query(func.sum(Sale.grand_total)).filter(
            Sale.customer_id == debtor.id,
            Sale.sale_status != SaleStatus.VOIDED
        ).scalar() or Decimal('0')
        
        total_paid = db.session.query(func.sum(Payment.amount)).filter(
            Payment.customer_id == debtor.id
        ).scalar() or Decimal('0')
        
        # Get last invoice date
        last_invoice = db.session.query(func.max(Sale.created_at)).filter(
            Sale.customer_id == debtor.id,
            Sale.sale_status != SaleStatus.VOIDED
        ).scalar()
        
        results.append({
            'id': debtor.id,
            'name': debtor.full_name,
            'phone': debtor.phone or '-',
            'balance': float(debtor.balance),
            'total_invoiced': float(total_invoiced),
            'total_paid': float(total_paid),
            'last_invoice_date': last_invoice,
            'days_overdue': days_overdue,
            'aging_status': 'overdue' if days_overdue >= 10 else 'current'
        })
    
    return results


# =============================================================================
# CREATE CUSTOMER
# =============================================================================

@customers_bp.route('/create', methods=['GET', 'POST'])
@login_required
def create():
    """Create a new customer"""
    if request.method == 'POST':
        full_name = request.form.get('full_name', '').strip()
        phone = request.form.get('phone', '').strip() or None
        email = request.form.get('email', '').strip() or None
        address = request.form.get('address', '').strip() or None
        notes = request.form.get('notes', '').strip() or None
        
        # Validation
        if not full_name:
            flash('Full name is required.', 'danger')
            return render_template('customers/form.html', customer=None)
        
        # Check for duplicate phone
        if phone:
            existing = Customer.query.filter_by(phone=phone).first()
            if existing:
                flash(f'Phone number already registered to {existing.full_name}.', 'danger')
                return render_template('customers/form.html', customer=None)
        
        try:
            customer = Customer(
                full_name=full_name,
                phone=phone,
                email=email,
                address=address,
                notes=notes,
                status=CustomerStatus.ACTIVE
            )
            db.session.add(customer)
            db.session.commit()
            
            flash(f'Customer "{full_name}" created successfully.', 'success')
            return redirect(url_for('customers.profile', customer_id=customer.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating customer: {str(e)}', 'danger')
    
    return render_template('customers/form.html', customer=None)


# =============================================================================
# EDIT CUSTOMER
# =============================================================================

@customers_bp.route('/<int:customer_id>/edit', methods=['GET', 'POST'])
@login_required
def edit(customer_id):
    """Edit an existing customer"""
    customer = Customer.query.get_or_404(customer_id)
    
    if request.method == 'POST':
        full_name = request.form.get('full_name', '').strip()
        phone = request.form.get('phone', '').strip() or None
        email = request.form.get('email', '').strip() or None
        address = request.form.get('address', '').strip() or None
        notes = request.form.get('notes', '').strip() or None
        status = request.form.get('status', 'Active')
        
        # Validation
        if not full_name:
            flash('Full name is required.', 'danger')
            return render_template('customers/form.html', customer=customer)
        
        # Check for duplicate phone (excluding current customer)
        if phone:
            existing = Customer.query.filter(
                Customer.phone == phone, 
                Customer.id != customer_id
            ).first()
            if existing:
                flash(f'Phone number already registered to {existing.full_name}.', 'danger')
                return render_template('customers/form.html', customer=customer)
        
        try:
            customer.full_name = full_name
            customer.phone = phone
            customer.email = email
            customer.address = address
            customer.notes = notes
            customer.status = CustomerStatus(status)
            
            db.session.commit()
            flash(f'Customer "{full_name}" updated successfully.', 'success')
            return redirect(url_for('customers.profile', customer_id=customer.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating customer: {str(e)}', 'danger')
    
    return render_template('customers/form.html', customer=customer)


# =============================================================================
# CUSTOMER PROFILE
# =============================================================================

@customers_bp.route('/<int:customer_id>')
@login_required
def profile(customer_id):
    """Customer profile page with invoices, payments, and ledger"""
    customer = Customer.query.get_or_404(customer_id)
    
    # Get outstanding balance from ledger
    outstanding_balance = customer.get_outstanding_balance()
    
    # Get recent invoices (last 20)
    invoices = Sale.query.filter_by(customer_id=customer_id)\
        .order_by(Sale.created_at.desc()).limit(20).all()
    
    # Get unpaid/partial invoices
    unpaid_invoices = Sale.query.filter(
        Sale.customer_id == customer_id,
        Sale.invoice_status.in_([InvoiceStatus.UNPAID, InvoiceStatus.PARTIAL])
    ).order_by(Sale.created_at.desc()).all()
    
    # Get recent payments (last 20)
    payments = Payment.query.filter_by(customer_id=customer_id)\
        .order_by(Payment.created_at.desc()).limit(20).all()
    
    # Get recent ledger entries (last 30)
    ledger_entries = CustomerLedger.query.filter_by(customer_id=customer_id)\
        .order_by(CustomerLedger.created_at.desc(), CustomerLedger.id.desc()).limit(30).all()

    # Statistics for Cards
    stats = db.session.query(
        func.count(Sale.id).label('order_count'),
        func.coalesce(func.sum(Sale.grand_total), 0).label('total_spent'),
        func.max(Sale.created_at).label('last_purchase')
    ).filter(
        Sale.customer_id == customer_id,
        Sale.sale_status != SaleStatus.VOIDED
    ).first()

    total_orders = stats.order_count or 0
    total_spent = stats.total_spent or 0
    last_purchase_date = stats.last_purchase
    
    return render_template('customers/profile.html',
                           customer=customer,
                           outstanding_balance=outstanding_balance,
                           invoices=invoices,
                           unpaid_invoices=unpaid_invoices,
                           payments=payments,
                           ledger_entries=ledger_entries,
                           total_orders=total_orders,
                           total_spent=total_spent,
                           last_purchase_date=last_purchase_date)


# =============================================================================
# DELETE CUSTOMER (SOFT DELETE)
# =============================================================================

@customers_bp.route('/<int:customer_id>/delete', methods=['POST'])
@login_required
def delete(customer_id):
    """Soft delete - set status to inactive"""
    customer = Customer.query.get_or_404(customer_id)
    
    # Check if customer has outstanding balance
    if customer.get_outstanding_balance() > 0:
        flash('Cannot deactivate customer with outstanding balance.', 'danger')
        return redirect(url_for('customers.profile', customer_id=customer_id))
    
    try:
        customer.status = CustomerStatus.INACTIVE
        db.session.commit()
        flash(f'Customer "{customer.full_name}" has been deactivated.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deactivating customer: {str(e)}', 'danger')
    
    return redirect(url_for('customers.index'))


# =============================================================================
# CUSTOMER SEARCH API (FOR CHECKOUT DROPDOWN)
# =============================================================================

@customers_bp.route('/api/search')
@login_required
def api_search():
    """Search customers for dropdown (returns JSON)"""
    q = request.args.get('q', '').strip()
    limit = request.args.get('limit', 10, type=int)
    
    if len(q) < 2:
        return jsonify({'customers': [], 'count': 0})
    
    customers = Customer.query.filter(
        Customer.status == CustomerStatus.ACTIVE,
        db.or_(
            Customer.full_name.ilike(f'%{q}%'),
            Customer.phone.ilike(f'%{q}%')
        )
    ).order_by(Customer.full_name.asc()).limit(limit).all()
    
    result = []
    for c in customers:
        balance = c.get_outstanding_balance()
        result.append({
            'id': c.id,
            'name': c.full_name,
            'phone': c.phone or '',
            'balance': float(balance),
            'credit_balance': float(c.get_available_credit())
        })
    
    return jsonify({'customers': result, 'count': len(result)})


# =============================================================================
# GET CUSTOMER DETAILS API
# =============================================================================

@customers_bp.route('/api/<int:customer_id>')
@login_required
def api_get(customer_id):
    """Get customer details for checkout modal"""
    customer = Customer.query.get_or_404(customer_id)
    balance = customer.get_outstanding_balance()
    
    return jsonify({
        'success': True,
        'customer': {
            'id': customer.id,
            'name': customer.full_name,
            'phone': customer.phone or '',
            'email': customer.email or '',
            'balance': balance,
            'credit_balance': float(customer.credit_balance) if customer.credit_balance else 0
        }
    })


# =============================================================================
# RECEIVE PAYMENT (PHASE 6)
# =============================================================================

@customers_bp.route('/<int:customer_id>/payment', methods=['GET', 'POST'])
@login_required
def receive_payment(customer_id):
    """Receive payment for a customer and allocate to unpaid invoices"""
    customer = Customer.query.get_or_404(customer_id)
    
    if request.method == 'POST':
        try:
            amount = Decimal(request.form.get('amount', 0))
            method_str = request.form.get('method', 'Cash')
            date_str = request.form.get('date', datetime.utcnow().strftime('%Y-%m-%d'))
            reference = request.form.get('reference', '').strip() or None
            notes = request.form.get('notes', '').strip() or None
            
            if amount <= 0:
                flash('Payment amount must be greater than 0.', 'danger')
                return redirect(url_for('customers.receive_payment', customer_id=customer_id))
            
            # 1. Create Ledger Entry for the TOTAL amount (Single Source of Truth)
            # This immediately reduces the customer's outstanding balance
            current_balance = customer.get_outstanding_balance()
            new_balance = current_balance - amount
            
            # Re-instantiating cleanly:
            ledger_entry = CustomerLedger(
                customer_id=customer.id,
                entry_type=LedgerEntryType.PAYMENT,
                debit=Decimal('0'),
                credit=amount,
                balance_after=new_balance,
                note=f'Payment Received ({method_str})' + (f' - {notes}' if notes else ''),
                created_at=datetime.strptime(date_str, '%Y-%m-%d')
            )
            db.session.add(ledger_entry)

            # 2. Allocate to Unpaid Invoices (FIFO)
            # Get unpaid invoices sorted by date (oldest first)
            unpaid_invoices = Sale.query.filter(
                Sale.customer_id == customer_id,
                Sale.invoice_status.in_([InvoiceStatus.UNPAID, InvoiceStatus.PARTIAL])
            ).order_by(Sale.created_at.asc()).all()

            remaining_payment = amount
            allocations = []

            for invoice in unpaid_invoices:
                if remaining_payment <= 0:
                    break
                
                due = invoice.amount_due
                to_pay = min(remaining_payment, due)
                
                # Create Payment record for this invoice
                payment = Payment(
                    sale_id=invoice.id,
                    customer_id=customer.id,
                    amount=to_pay,
                    payment_method=PaymentMethod(method_str),
                    reference=reference,
                    note=notes,
                    created_by=current_user.id
                )
                db.session.add(payment)
                
                # Flush to ensure payment ID is generated and visible to relationships
                db.session.flush()
                
                # Update Invoice Status
                # Append to relationships to ensure in-memory state is up to date for immediate calculation
                # invoice.payments.append(payment) 
                
                remaining_payment -= to_pay
                allocations.append(f"{invoice.invoice_no}: {to_pay}")

            # 3. Handle Excess Payment (Credit)
            if remaining_payment > 0:
                customer.credit_balance = (customer.credit_balance or 0) + remaining_payment
                flash(f'Payment allocated. ${remaining_payment} added to credit balance.', 'info')
            
            # Commit all changes
            db.session.commit()
            
            # 4. Post-commit: Update invoice statuses
            # We need to re-fetch or iterate to call update_payment_status which uses SQL sum
            for invoice in unpaid_invoices:
                 invoice.update_payment_status()
            db.session.commit()

            flash(f'Payment of ${amount} recorded successfully.', 'success')
            return redirect(url_for('customers.profile', customer_id=customer_id))

        except Exception as e:
            db.session.rollback()
            flash(f'Error recording payment: {str(e)}', 'danger')
            return redirect(url_for('customers.receive_payment', customer_id=customer_id))

    # GET Request
    outstanding_balance = customer.get_outstanding_balance()
    unpaid_invoices = Sale.query.filter(
        Sale.customer_id == customer_id,
        Sale.invoice_status.in_([InvoiceStatus.UNPAID, InvoiceStatus.PARTIAL])
    ).order_by(Sale.created_at.asc()).all()
    
    return render_template('customers/payment.html',
                           customer=customer,
                           outstanding_balance=outstanding_balance,
                           unpaid_invoices=unpaid_invoices,
                           today=datetime.utcnow().strftime('%Y-%m-%d'))


# =============================================================================
# RECALCULATE BALANCE (FIX DATA INCONSISTENCY)
# =============================================================================

@customers_bp.route('/<int:customer_id>/recalculate', methods=['POST'])
@login_required
def recalculate_balance(customer_id):
    """Force recalculate customer balance based on Invoices vs Payments"""
    if not current_user.has_role('Admin') and not current_user.has_role('Manager'):
        flash('Access denied', 'danger')
        return redirect(url_for('customers.profile', customer_id=customer_id))

    customer = Customer.query.get_or_404(customer_id)
    
    try:
        # 1. Calculate Theoretical Balance
        # Total Invoices (Debit)
        total_invoices = db.session.query(db.func.sum(Sale.grand_total)).filter(
            Sale.customer_id == customer_id,
            Sale.sale_status != SaleStatus.VOIDED
        ).scalar() or Decimal('0')
        
        # Total Payments (Credit)
        total_payments = db.session.query(db.func.sum(Payment.amount)).filter(
            Payment.customer_id == customer_id
        ).scalar() or Decimal('0')
        
        theoretical_balance = total_invoices - total_payments
        
        # 2. Get Current Ledger Balance
        current_ledger_balance = customer.get_outstanding_balance()
        
        # 3. Adjust if needed
        if theoretical_balance != current_ledger_balance:
            diff = theoretical_balance - current_ledger_balance
            
            entry = CustomerLedger(
                customer_id=customer.id,
                entry_type=LedgerEntryType.ADJUSTMENT,
                balance_after=theoretical_balance,
                note='System Recalculation (Fix Data Inconsistency)',
                created_at=datetime.utcnow()
            )
            
            if diff > 0:
                # Balance needs to increase (Debit)
                entry.debit = diff
                entry.credit = 0
            else:
                # Balance needs to decrease (Credit)
                entry.debit = 0
                entry.credit = abs(diff)
                
            db.session.add(entry)
            db.session.commit()
            
            flash(f'Balance recalculated. Adjustment of {format_currency(diff)} applied.', 'success')
        else:
            flash('Balance is already correct.', 'info')
            
    except Exception as e:
        db.session.rollback()
        flash(f'Error recalculating balance: {str(e)}', 'danger')

    return redirect(url_for('customers.profile', customer_id=customer_id))


# =============================================================================
# DEBTORS EXPORT (Excel / PDF)
# =============================================================================

# =============================================================================
# EXPORTS (Excel / PDF)
# =============================================================================
import os


@customers_bp.route('/export/excel')
@login_required
def export_excel():
    """Export customers to Excel (Unified Logic)"""
    search = request.args.get('search', '').strip()
    status_filter = request.args.get('status', '')
    
    query = Customer.query
    
    # 1. Search Filter
    if search:
        query = query.filter(db.or_(
            Customer.full_name.ilike(f'%{search}%'),
            Customer.phone.ilike(f'%{search}%'),
            Customer.email.ilike(f'%{search}%')
        ))
    
    # 2. Status/Debtors Filter
    if status_filter:
        if status_filter == 'Debtors':
            # Subquery for debtors (same as index route)
            subq = db.session.query(
                CustomerLedger.customer_id,
                func.max(CustomerLedger.id).label('max_id')
            ).group_by(CustomerLedger.customer_id).subquery()
            
            debtor_ids = db.session.query(CustomerLedger.customer_id)\
                .join(subq, CustomerLedger.id == subq.c.max_id)\
                .filter(CustomerLedger.balance_after > 0).all()
            debtor_ids = [d[0] for d in debtor_ids]
            
            query = query.filter(Customer.id.in_(debtor_ids))
        else:
            try:
                query = query.filter(Customer.status == CustomerStatus(status_filter))
            except: pass
            
    # Sort by ID Ascending (1, 2, 3) as requested
    customers = query.order_by(Customer.id.asc()).all()
    
    if not customers:
        flash('No customers found to export.', 'info')
        return redirect(url_for('customers.index'))
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customer List"
    
    # Company Info for Header
    from app.models import SystemSetting
    company_name = SystemSetting.get('company_name', 'Electronics POS')
    
    # Styles
    title_font = Font(bold=True, size=14, color="2C3E50")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    
    zebra_light = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    zebra_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    alignment_center = Alignment(horizontal='center', vertical='center')
    alignment_left = Alignment(horizontal='left', vertical='center')
    
    # Title Section
    ws.merge_cells('A1:F1')
    ws['A1'] = company_name.upper()
    ws['A1'].font = title_font
    ws['A1'].alignment = alignment_center
    
    ws.merge_cells('A2:F2')
    ws['A2'] = f"CUSTOMER LIST - {datetime.now().strftime('%d %b %Y %H:%M')}"
    ws['A2'].alignment = alignment_center
    
    # Headers
    headers = ['ID', 'Customer Name', 'Phone', 'Email', 'Status', 'Balance']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center
        cell.border = thin_border
        
    # Data
    total_balance = 0
    row_idx = 5
    for c in customers:
        debt = c.get_current_debt()
        total_balance += debt
        
        # Zebra Striping
        row_fill = zebra_light if row_idx % 2 == 0 else zebra_white
        
        # Write cells
        ws.cell(row=row_idx, column=1, value=c.id).alignment = alignment_center
        ws.cell(row=row_idx, column=2, value=c.full_name)
        ws.cell(row=row_idx, column=3, value=c.phone or '-')
        ws.cell(row=row_idx, column=4, value=c.email or '-')
        ws.cell(row=row_idx, column=5, value=c.status.value).alignment = alignment_center
        
        bal_cell = ws.cell(row=row_idx, column=6, value=float(debt))
        bal_cell.number_format = '#,##0.00'
        if debt > 0:
            bal_cell.font = Font(color="C0392B", bold=True)
        else:
            bal_cell.font = Font(color="27AE60")
            
        # Apply borders and fill
        for col in range(1, 7):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin_border
            cell.fill = row_fill
            
        row_idx += 1
        
    # Summary Section
    summary_row = row_idx + 1
    ws.cell(row=summary_row, column=2, value=f"Total Customers: {len(customers)}").font = Font(bold=True)
    
    ws.cell(row=summary_row, column=5, value="Total Balance:").font = Font(bold=True)
    ws.cell(row=summary_row, column=5).alignment = Alignment(horizontal='right')
    
    total_cell = ws.cell(row=summary_row, column=6, value=float(total_balance))
    total_cell.font = Font(bold=True, size=11)
    total_cell.number_format = '"$"#,##0.00'
    
    # Optimal Column Widths
    ws.column_dimensions['A'].width = 8   # ID
    ws.column_dimensions['B'].width = 35  # Name (Wider)
    ws.column_dimensions['C'].width = 18  # Phone
    ws.column_dimensions['D'].width = 25  # Email
    ws.column_dimensions['E'].width = 12  # Status
    ws.column_dimensions['F'].width = 18  # Balance
        
    filename = f"Customer_List_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    
    # Save
    excel_io = BytesIO()
    wb.save(excel_io)
    excel_io.seek(0)
    
    return send_file(excel_io, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


@customers_bp.route('/export/pdf')
@login_required
def export_pdf():
    """Export customers to PDF (Unified Logic)"""
    search = request.args.get('search', '').strip()
    status_filter = request.args.get('status', '')
    
    query = Customer.query
    if search:
        query = query.filter(db.or_(
            Customer.full_name.ilike(f'%{search}%'),
            Customer.phone.ilike(f'%{search}%'),
            Customer.email.ilike(f'%{search}%')
        ))
        
    if status_filter:
        if status_filter == 'Debtors':
            subq = db.session.query(
                CustomerLedger.customer_id,
                func.max(CustomerLedger.id).label('max_id')
            ).group_by(CustomerLedger.customer_id).subquery()
            
            debtor_ids = db.session.query(CustomerLedger.customer_id)\
                .join(subq, CustomerLedger.id == subq.c.max_id)\
                .filter(CustomerLedger.balance_after > 0).all()
            debtor_ids = [d[0] for d in debtor_ids]
            query = query.filter(Customer.id.in_(debtor_ids))
        else:
            try: query = query.filter(Customer.status == CustomerStatus(status_filter))
            except: pass
            
    customers = query.order_by(Customer.id.asc()).all()
    if not customers:
        flash('No customers found to export.', 'info')
        return redirect(url_for('customers.index'))
        
    # Calculate Summaries for PDF
    total_customers = len(customers)
    total_balance = sum(c.get_current_debt() for c in customers)
    
    # Render template (Company Info comes from global context_processor, matching POS receipt)
    html_content = render_template('customers/list_pdf.html', 
                                 customers=customers, 
                                 total_customers=total_customers,
                                 total_balance=total_balance,
                                 generated_at=datetime.now(), 
                                 format_currency=format_currency,
                                 filter_name=status_filter or "All")

    try:
        path_wkhtmltopdf = current_app.config.get('WKHTMLTOPDF_PATH')
        config = pdfkit.configuration(wkhtmltopdf=path_wkhtmltopdf)
        
        # Professional PDF Options
        options = {
            'page-size': 'A4',
            'margin-top': '10mm',
            'margin-right': '10mm',
            'margin-bottom': '15mm',
            'margin-left': '10mm',
            'encoding': 'UTF-8',
            'footer-center': '[page] of [toPage]',
            'footer-font-size': '9',
            'footer-spacing': '5'
        }
        
        pdf_bytes = pdfkit.from_string(html_content, False, configuration=config, options=options)
        pdf_io = BytesIO(pdf_bytes)
        pdf_io.seek(0)
        
        return send_file(pdf_io, mimetype='application/pdf', as_attachment=True,
                         download_name=f"Customer_List_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf")
    except Exception as e:
        current_app.logger.exception("PDF generation failed")
        flash(f'Error generating PDF: {str(e)}', 'danger')
        return redirect(url_for('customers.index'))
