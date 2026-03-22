from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, abort
from flask_login import login_required, current_user
from app.models import Product, Category, SaleItem
from app import db
from app.forms import ProductForm
from app.services.audit_service import AuditService
from app.services.barcode_service import BarcodeService
from app.services.label_service import LabelService
from sqlalchemy import or_, desc
from werkzeug.utils import secure_filename
from collections import defaultdict
from decimal import Decimal
import json
import uuid
import os
import pdfkit
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from flask import send_file, Response, current_app
from app.models import SystemSetting
from app.utils import format_currency


products_bp = Blueprint('products', __name__, url_prefix='/products')

# ---------------------------------------------------------------------------
# Helper: permission guard
# ---------------------------------------------------------------------------

def _require_manager():
    """Return True if current_user is Admin or Manager, flash + False otherwise."""
    if current_user.has_role('Admin') or current_user.has_role('Manager'):
        return True
    flash('Access denied', 'danger')
    return False


# ---------------------------------------------------------------------------
# Product list
# ---------------------------------------------------------------------------

@products_bp.route('/')
@login_required
def index():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '')
    category_id = request.args.get('category_id', type=int)

    query = Product.query

    if search:
        query = query.filter(or_(
            Product.name.contains(search),
            Product.sku.contains(search),
            Product.barcode.contains(search)
        ))

    if category_id:
        query = query.filter(Product.category_id == category_id)

    products = query.order_by(desc(Product.created_at)).paginate(
        page=page, per_page=10, error_out=False
    )

    categories = Category.query.all()

    return render_template('products/index.html',
                           products=products,
                           categories=categories,
                           search=search,
                           category_id=category_id)


# ---------------------------------------------------------------------------
# Add product
# ---------------------------------------------------------------------------

@products_bp.route('/add', methods=['GET', 'POST'])
@login_required
def add():
    if not _require_manager():
        return redirect(url_for('products.index'))

    form = ProductForm()
    form.category_id.choices = [(c.id, c.name) for c in Category.query.all()]

    if form.validate_on_submit():
        try:
            barcode_input = (form.barcode.data or '').strip()

            # Uniqueness check for manually provided barcode
            if barcode_input and not BarcodeService.is_barcode_unique(barcode_input):
                flash('Barcode already assigned to another product.', 'danger')
                return render_template('products/add.html', form=form)

            product = Product(
                name=form.name.data,
                category_id=form.category_id.data,
                sku=form.sku.data,
                barcode=barcode_input or None,
                cost_price=form.cost_price.data,
                selling_price=form.selling_price.data,
                wholesale_price=form.wholesale_price.data,
                min_wholesale_qty=form.min_wholesale_qty.data,
                allow_wholesale=form.allow_wholesale.data,
                quantity_in_stock=form.quantity.data,
                low_stock_threshold=form.low_stock_threshold.data,
                description=form.description.data
            )

            # Classify barcode source when user provided one
            if barcode_input:
                product.barcode_type = BarcodeService.detect_type(barcode_input)
                product.barcode_source = 'manual'

            # Handle image upload before flush
            if form.image.data and form.image.data.filename:
                product.image_filename = _save_product_image(form.image.data)

            db.session.add(product)
            db.session.flush()  # get product.id before barcode generation

            # Auto-generate barcode if none was provided
            barcode_generated = BarcodeService.ensure_product_barcode(product)

            db.session.commit()

            audit_details = {'name': product.name, 'sku': product.sku, 'barcode': product.barcode}
            if barcode_generated:
                audit_details['barcode_action'] = 'auto_generated'
            AuditService.log_action(
                action='CREATE_PRODUCT',
                target_type='Product',
                target_id=product.id,
                details=audit_details
            )
            if barcode_generated:
                AuditService.log_action(
                    action='BARCODE_GENERATED',
                    target_type='Product',
                    target_id=product.id,
                    details={
                        'barcode': product.barcode,
                        'barcode_type': product.barcode_type,
                        'source': 'auto_on_create',
                    }
                )

            flash(f'Product added. Barcode: {product.barcode}', 'success')
            return redirect(url_for('products.index'))

        except Exception as e:
            db.session.rollback()
            import traceback
            current_app.logger.error(f"Error adding product: {e}\n{traceback.format_exc()}")
            flash(f'Error adding product: {str(e)}', 'danger')

    return render_template('products/add.html', form=form)


# ---------------------------------------------------------------------------
# Edit product
# ---------------------------------------------------------------------------

@products_bp.route('/<int:product_id>/edit', methods=['GET', 'POST'])
@login_required
def edit(product_id):
    if not _require_manager():
        return redirect(url_for('products.index'))

    product = Product.query.get_or_404(product_id)
    form = ProductForm(obj=product)

    if request.method == 'GET':
        form.quantity.data = product.quantity_in_stock

    form.category_id.choices = [(c.id, c.name) for c in Category.query.all()]

    if form.validate_on_submit():
        barcode_input = (form.barcode.data or '').strip()
        old_barcode = product.barcode

        # Uniqueness check when barcode changes
        if barcode_input and barcode_input != old_barcode:
            if not BarcodeService.is_barcode_unique(barcode_input, exclude_product_id=product.id):
                flash('Barcode already assigned to another product.', 'danger')
                return render_template('products/edit.html', form=form, product=product)

        product.name = form.name.data
        product.category_id = form.category_id.data
        product.sku = form.sku.data
        product.cost_price = form.cost_price.data
        product.selling_price = form.selling_price.data
        product.wholesale_price = form.wholesale_price.data
        product.min_wholesale_qty = form.min_wholesale_qty.data
        product.allow_wholesale = form.allow_wholesale.data
        product.quantity_in_stock = form.quantity.data
        product.low_stock_threshold = form.low_stock_threshold.data
        product.description = form.description.data

        barcode_changed = False
        if barcode_input and barcode_input != old_barcode:
            BarcodeService.apply_manual_barcode(product, barcode_input)
            barcode_changed = True
        elif not barcode_input and not product.barcode:
            # No barcode at all — generate one
            BarcodeService.ensure_product_barcode(product)

        # Image handling
        if form.image.data and form.image.data.filename:
            if product.image_filename:
                _delete_product_image(product.image_filename)
            product.image_filename = _save_product_image(form.image.data)

        if request.form.get('remove_image') == '1' and product.image_filename:
            _delete_product_image(product.image_filename)
            product.image_filename = None

        db.session.commit()

        audit_details = {'name': product.name, 'sku': product.sku}
        if barcode_changed:
            audit_details['barcode_old'] = old_barcode
            audit_details['barcode_new'] = product.barcode
            audit_details['barcode_source'] = product.barcode_source
        AuditService.log_action(
            action='UPDATE_PRODUCT',
            target_type='Product',
            target_id=product.id,
            details=audit_details
        )

        flash('Product updated successfully', 'success')
        return redirect(url_for('products.index'))

    return render_template('products/edit.html', form=form, product=product)


# ---------------------------------------------------------------------------
# Delete product
# ---------------------------------------------------------------------------

@products_bp.route('/<int:product_id>/delete', methods=['POST'])
@login_required
def delete(product_id):
    if not current_user.has_role('Admin'):
        flash('Access denied', 'danger')
        return redirect(url_for('products.index'))

    product = Product.query.get_or_404(product_id)

    if SaleItem.query.filter_by(product_id=product_id).first():
        flash('Cannot delete product with existing sales', 'danger')
        return redirect(url_for('products.index'))

    if product.image_filename:
        _delete_product_image(product.image_filename)

    db.session.delete(product)
    db.session.commit()

    AuditService.log_action(
        action='DELETE_PRODUCT',
        target_type='Product',
        target_id=product.id,
        details={'name': product.name, 'sku': product.sku}
    )

    flash('Product deleted successfully', 'success')
    return redirect(url_for('products.index'))


# ---------------------------------------------------------------------------
# Search
# ---------------------------------------------------------------------------

@products_bp.route('/search')
@login_required
def search():
    query = request.args.get('q', '')
    if not query:
        return jsonify([])

    products = Product.query.filter(
        or_(
            Product.name.contains(query),
            Product.sku.contains(query),
            Product.barcode.contains(query)
        )
    ).limit(10).all()

    results = []
    for product in products:
        results.append({
            'id': product.id,
            'name': product.name,
            'sku': product.sku,
            'barcode': product.barcode,
            'price': float(product.selling_price),
            'quantity': product.quantity_in_stock,
            'category': product.category.name if product.category else ''
        })

    return jsonify(results)


# ---------------------------------------------------------------------------
# Low stock
# ---------------------------------------------------------------------------

@products_bp.route('/low-stock')
@login_required
def low_stock():
    if not _require_manager():
        return redirect(url_for('products.index'))

    products = Product.query.filter(Product.quantity_in_stock <= Product.low_stock_threshold).all()
    return render_template('products/low_stock.html', products=products)


@products_bp.route('/low-stock/pdf')
@login_required
def low_stock_pdf():
    if not _require_manager():
        return redirect(url_for('products.index'))

    products = Product.query.filter(
        Product.quantity_in_stock <= Product.low_stock_threshold
    ).order_by(Product.quantity_in_stock.asc(), Product.name.asc()).all()

    # Company info
    company_name    = SystemSetting.get('company_name', 'Electronics Store POS')
    company_address = SystemSetting.get('company_address', '')
    company_phone   = SystemSetting.get('company_phone', '')
    company_logo    = SystemSetting.get('company_logo', '')

    # Summary stats
    critical_items = [p for p in products if p.quantity_in_stock == 0]
    low_items      = [p for p in products if p.quantity_in_stock > 0]
    total_needed   = sum(max(p.low_stock_threshold - p.quantity_in_stock, 0) for p in products)
    retail_at_risk = sum(
        (p.selling_price or Decimal('0')) * max(p.low_stock_threshold - p.quantity_in_stock, 0)
        for p in products
    )

    html = render_template(
        'products/low_stock_pdf.html',
        products=products,
        critical_items=critical_items,
        low_items=low_items,
        total_needed=total_needed,
        retail_at_risk=retail_at_risk,
        generated_at=datetime.now().strftime('%B %d, %Y'),
        generated_time=datetime.now().strftime('%H:%M'),
        company_name=company_name,
        company_address=company_address,
        company_phone=company_phone,
        company_logo=company_logo,
        format_currency=format_currency,
    )

    try:
        path_wkhtmltopdf = current_app.config.get('WKHTMLTOPDF_PATH')
        config  = pdfkit.configuration(wkhtmltopdf=path_wkhtmltopdf)
        options = {
            'page-size':               'A4',
            'orientation':             'Portrait',
            'margin-top':              '15mm',
            'margin-bottom':           '18mm',
            'margin-left':             '15mm',
            'margin-right':            '15mm',
            'encoding':                'UTF-8',
            'no-outline':              None,
            'enable-local-file-access': None,
            'dpi':                     '96',
            'zoom':                    '1',
            'print-media-type':        None,
            'disable-smart-shrinking': None,
            'minimum-font-size':       '10',
            'footer-left':             company_name,
            'footer-center':           f"Low Stock Alert \u2014 {datetime.now().strftime('%B %d, %Y')}",
            'footer-right':            'Page [page] of [topage]',
            'footer-font-name':        'Arial',
            'footer-font-size':        '8',
            'footer-spacing':          '4',
        }
        pdf_bytes = pdfkit.from_string(html, False, configuration=config, options=options)
    except Exception as e:
        current_app.logger.exception("pdfkit failed to generate low stock PDF")
        flash(f"PDF generation failed: {str(e)}", "danger")
        return redirect(url_for('products.low_stock'))

    pdf_io = BytesIO(pdf_bytes)
    pdf_io.seek(0)
    filename = f"low_stock_alert_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return send_file(pdf_io, mimetype='application/pdf',
                     as_attachment=True, download_name=filename)


# ---------------------------------------------------------------------------
# Categories
# ---------------------------------------------------------------------------

@products_bp.route('/categories', methods=['GET', 'POST'])
@login_required
def categories():
    if not _require_manager():
        return redirect(url_for('products.index'))

    if request.method == 'POST':
        name = request.form.get('name')
        description = request.form.get('description')

        if Category.query.filter_by(name=name).first():
            flash('Category with this name already exists.', 'warning')
        else:
            try:
                category = Category(name=name, description=description)
                db.session.add(category)
                db.session.commit()
                AuditService.log_action(
                    action='CREATE_CATEGORY',
                    target_type='Category',
                    target_id=category.id,
                    details={'name': category.name}
                )
                flash('Category added successfully.', 'success')
            except Exception as e:
                db.session.rollback()
                flash(f'Error adding category: {str(e)}', 'danger')
        return redirect(url_for('products.categories'))

    categories = Category.query.order_by(Category.name).all()
    return render_template('products/categories.html', categories=categories)


@products_bp.route('/categories/<int:id>/edit', methods=['POST'])
@login_required
def edit_category(id):
    if not _require_manager():
        return redirect(url_for('products.index'))

    category = Category.query.get_or_404(id)
    try:
        category.name = request.form.get('name')
        category.description = request.form.get('description')
        db.session.commit()
        AuditService.log_action(
            action='UPDATE_CATEGORY',
            target_type='Category',
            target_id=category.id,
            details={'name': category.name}
        )
        flash('Category updated successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error updating category: {str(e)}', 'danger')

    return redirect(url_for('products.categories'))


@products_bp.route('/categories/<int:id>/delete', methods=['POST'])
@login_required
def delete_category(id):
    if not current_user.has_role('Admin'):
        flash('Access denied', 'danger')
        return redirect(url_for('products.index'))

    category = Category.query.get_or_404(id)
    if category.products:
        flash('Cannot delete category because it contains products.', 'warning')
        return redirect(url_for('products.categories'))

    try:
        db.session.delete(category)
        db.session.commit()
        AuditService.log_action(
            action='DELETE_CATEGORY',
            target_type='Category',
            target_id=category.id,
            details={'name': category.name}
        )
        flash('Category deleted successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting category: {str(e)}', 'danger')

    return redirect(url_for('products.categories'))


# ===========================================================================
# BARCODE ENDPOINTS
# ===========================================================================

@products_bp.route('/<int:product_id>/barcode/regenerate', methods=['POST'])
@login_required
def regenerate_barcode(product_id):
    """Admin/Manager only: generate a fresh barcode, discarding the current one."""
    if not (current_user.has_role('Admin') or current_user.has_role('Manager')):
        return jsonify({'success': False, 'message': 'Access denied'}), 403

    product = Product.query.get_or_404(product_id)
    old_barcode = product.barcode

    try:
        product.barcode = None
        product.barcode_type = None
        product.barcode_source = None
        db.session.flush()

        BarcodeService.ensure_product_barcode(product)
        db.session.commit()

        AuditService.log_action(
            action='BARCODE_REGENERATED',
            target_type='Product',
            target_id=product.id,
            details={
                'old_barcode': old_barcode,
                'new_barcode': product.barcode,
                'barcode_type': product.barcode_type,
            }
        )

        return jsonify({
            'success': True,
            'barcode': product.barcode,
            'barcode_type': product.barcode_type,
            'barcode_source': product.barcode_source,
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"regenerate_barcode error for product {product_id}: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


# ---------------------------------------------------------------------------
# Barcode image endpoint — for sidebar preview
# ---------------------------------------------------------------------------

@products_bp.route('/<int:product_id>/barcode.png')
@login_required
def barcode_image(product_id):
    """Stream a PNG barcode image for the sidebar preview panel."""
    product = Product.query.get_or_404(product_id)
    if not product.barcode:
        abort(404)
    try:
        img_bytes = BarcodeService.generate_barcode_image(
            product.barcode,
            product.barcode_type or 'Code128',
            options={
                'module_width':  0.35,
                'module_height': 12.0,
                'quiet_zone':    4.0,
                'font_size':     9,
                'text_distance': 3.0,
                'background':    'white',
                'foreground':    'black',
                'write_text':    True,
                'dpi':           150,
            },
        )
        if not img_bytes:
            abort(404)
        return Response(
            img_bytes,
            mimetype='image/png',
            headers={'Cache-Control': 'no-cache, must-revalidate'},
        )
    except Exception as exc:
        current_app.logger.warning('barcode_image error for product %s: %s', product_id, exc)
        abort(404)


# ---------------------------------------------------------------------------
# Single-label print
# ---------------------------------------------------------------------------

@products_bp.route('/<int:product_id>/label')
@login_required
def print_label(product_id):
    """Stream a PDF label for a single product.  ?copies=N&template=standard"""
    product = Product.query.get_or_404(product_id)
    copies   = max(1, min(request.args.get('copies',   1,          type=int), 100))
    template = request.args.get('template', 'standard')

    if not product.barcode:
        if current_user.has_role('Admin') or current_user.has_role('Manager'):
            try:
                BarcodeService.ensure_product_barcode(product)
                db.session.commit()
                AuditService.log_action(
                    action='BARCODE_GENERATED',
                    target_type='Product',
                    target_id=product.id,
                    details={'barcode': product.barcode, 'source': 'auto_on_label_print'}
                )
            except Exception as e:
                db.session.rollback()
                flash(f'Could not generate barcode: {e}', 'danger')
                return redirect(url_for('products.edit', product_id=product_id))
        else:
            flash('This product has no barcode. Ask a manager to assign one.', 'warning')
            return redirect(url_for('products.index'))

    try:
        pdf_bytes = LabelService.generate_single_label_pdf(product, copies=copies, template=template)
    except Exception as e:
        current_app.logger.error(f"print_label PDF error for product {product_id}: {e}")
        flash(f'Label PDF generation failed: {e}', 'danger')
        return redirect(url_for('products.edit', product_id=product_id))

    AuditService.log_action(
        action='LABEL_PRINTED',
        target_type='Product',
        target_id=product.id,
        details={'copies': copies, 'barcode': product.barcode}
    )

    pdf_io = BytesIO(pdf_bytes)
    pdf_io.seek(0)
    filename = f"label_{product.sku}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return send_file(pdf_io, mimetype='application/pdf',
                     as_attachment=False, download_name=filename)


# ---------------------------------------------------------------------------
# Batch label print
# ---------------------------------------------------------------------------

@products_bp.route('/labels/print', methods=['GET', 'POST'])
@login_required
def batch_print_labels():
    """GET: batch selection page.  POST: generate and stream label PDF."""
    if request.method == 'POST':
        if request.is_json:
            data = request.get_json()
            items_raw = data.get('items', [])
            output_format  = data.get('output_format', 'pdf')
            label_template = data.get('label_template', 'standard')
        else:
            product_ids    = request.form.getlist('product_ids')
            output_format  = request.form.get('output_format', 'pdf')
            label_template = request.form.get('label_template', 'standard')
            items_raw = [
                {
                    'product_id': int(pid),
                    'copies': int(request.form.get(f'copies_{pid}', 1))
                }
                for pid in product_ids if pid
            ]

        if not items_raw:
            flash('No products selected.', 'warning')
            return redirect(url_for('products.batch_print_labels'))

        label_items = []
        for item in items_raw:
            product = Product.query.get(item.get('product_id'))
            if not product:
                continue
            if not product.barcode and (current_user.has_role('Admin') or current_user.has_role('Manager')):
                try:
                    BarcodeService.ensure_product_barcode(product)
                    db.session.commit()
                except Exception:
                    db.session.rollback()
            label_items.append({
                'product': product,
                'copies':  max(1, min(int(item.get('copies', 1)), 100))
            })

        if not label_items:
            flash('No valid products to print.', 'warning')
            return redirect(url_for('products.batch_print_labels'))

        AuditService.log_action(
            action='BATCH_LABELS_PRINTED',
            target_type='Product',
            details={
                'product_count':   len(label_items),
                'total_copies':    sum(i['copies'] for i in label_items),
                'output_format':   output_format,
                'label_template':  label_template,
                'products':        [i['product'].id for i in label_items],
            }
        )

        ts = datetime.now().strftime('%Y%m%d_%H%M')

        # ZPL output — for direct thermal label printers
        if output_format == 'zpl':
            try:
                zpl_str = LabelService.generate_labels_zpl(label_items)
            except Exception as e:
                current_app.logger.error(f"batch_print_labels ZPL error: {e}")
                flash(f'ZPL generation failed: {e}', 'danger')
                return redirect(url_for('products.batch_print_labels'))
            zpl_io = BytesIO(zpl_str.encode('utf-8'))
            zpl_io.seek(0)
            return send_file(zpl_io, mimetype='text/plain',
                             as_attachment=True, download_name=f'labels_{ts}.zpl')

        # PDF output (default)
        try:
            pdf_bytes = LabelService.generate_labels_pdf(label_items, template=label_template)
        except Exception as e:
            current_app.logger.error(f"batch_print_labels PDF error: {e}")
            flash(f'Label PDF generation failed: {e}', 'danger')
            return redirect(url_for('products.batch_print_labels'))

        pdf_io = BytesIO(pdf_bytes)
        pdf_io.seek(0)
        return send_file(pdf_io, mimetype='application/pdf',
                         as_attachment=False, download_name=f'labels_{ts}.pdf')

    # GET
    search = request.args.get('search', '')
    category_id = request.args.get('category_id', type=int)

    query = Product.query.filter_by(is_active=True)
    if search:
        query = query.filter(or_(
            Product.name.contains(search),
            Product.sku.contains(search),
            Product.barcode.contains(search)
        ))
    if category_id:
        query = query.filter(Product.category_id == category_id)

    products_list = query.order_by(Product.name).all()
    categories = Category.query.order_by(Category.name).all()

    return render_template('products/batch_labels.html',
                           products=products_list,
                           categories=categories,
                           search=search,
                           category_id=category_id)


# ===========================================================================
# EXPORT ROUTES
# ===========================================================================

@products_bp.route('/export/pdf')
@login_required
def export_pdf():
    if not _require_manager():
        return redirect(url_for('products.index'))

    search = request.args.get('search', '')
    category_id = request.args.get('category_id', type=int)

    query = Product.query
    if search:
        query = query.filter(or_(
            Product.name.contains(search),
            Product.sku.contains(search),
            Product.barcode.contains(search)
        ))
    if category_id:
        query = query.filter(Product.category_id == category_id)

    products = query.order_by(Product.category_id, Product.name).all()

    category_filter_name = None
    if category_id:
        cat = Category.query.get(category_id)
        if cat:
            category_filter_name = cat.name

    # --- Company info ---
    company_name    = SystemSetting.get('company_name', 'Electronics Store POS')
    company_address = SystemSetting.get('company_address', '')
    company_phone   = SystemSetting.get('company_phone', '')
    company_logo    = SystemSetting.get('company_logo', '')

    # --- Summary stats ---
    total_products   = len(products)
    total_units      = sum(p.quantity_in_stock for p in products)
    low_stock_items  = [p for p in products if p.is_low_stock]
    ok_items         = [p for p in products if not p.is_low_stock]

    wholesale_value  = sum((p.cost_price or Decimal('0')) * p.quantity_in_stock for p in products)
    retail_value     = sum((p.selling_price or Decimal('0')) * p.quantity_in_stock for p in products)
    potential_margin = retail_value - wholesale_value

    # --- Category breakdown ---
    cat_map = defaultdict(lambda: {'products': 0, 'units': 0, 'wholesale': Decimal('0'), 'retail': Decimal('0')})
    for p in products:
        cat_name = p.category.name if p.category else 'Uncategorized'
        cat_map[cat_name]['products'] += 1
        cat_map[cat_name]['units']    += p.quantity_in_stock
        cat_map[cat_name]['wholesale'] += (p.cost_price or Decimal('0')) * p.quantity_in_stock
        cat_map[cat_name]['retail']    += (p.selling_price or Decimal('0')) * p.quantity_in_stock
    category_breakdown = sorted(cat_map.items(), key=lambda x: x[1]['retail'], reverse=True)

    html = render_template(
        'products/pdf_export.html',
        products=products,
        low_stock_items=low_stock_items,
        generated_at=datetime.now().strftime('%B %d, %Y'),
        category_filter=category_filter_name,
        search_query=search,
        format_currency=format_currency,
        # Company
        company_name=company_name,
        company_address=company_address,
        company_phone=company_phone,
        company_logo=company_logo,
        # Summary
        total_products=total_products,
        total_units=total_units,
        low_stock_count=len(low_stock_items),
        ok_count=len(ok_items),
        wholesale_value=wholesale_value,
        retail_value=retail_value,
        potential_margin=potential_margin,
        # Category data
        category_breakdown=category_breakdown,
    )

    try:
        path_wkhtmltopdf = current_app.config.get('WKHTMLTOPDF_PATH')
        config = pdfkit.configuration(wkhtmltopdf=path_wkhtmltopdf)
        options = {
            'page-size':               'A4',
            'orientation':             'Portrait',
            'margin-top':              '15mm',
            'margin-bottom':           '18mm',
            'margin-left':             '15mm',
            'margin-right':            '15mm',
            'encoding':                'UTF-8',
            'no-outline':              None,
            'enable-local-file-access': None,
            'dpi':                     '96',
            'zoom':                    '1',
            'print-media-type':        None,
            'disable-smart-shrinking': None,
            'minimum-font-size':       '10',
            'footer-left':             company_name,
            'footer-center':           f"Inventory Report \u2014 {datetime.now().strftime('%B %d, %Y')}",
            'footer-right':            'Page [page] of [topage]',
            'footer-font-name':        'Arial',
            'footer-font-size':        '8',
            'footer-spacing':          '4',
        }
        pdf_bytes = pdfkit.from_string(html, False, configuration=config, options=options)
    except Exception as e:
        current_app.logger.exception("pdfkit failed to generate inventory PDF")
        flash(f"PDF generation failed: {str(e)}", "danger")
        return redirect(url_for('products.index'))

    pdf_io = BytesIO(pdf_bytes)
    pdf_io.seek(0)
    filename = f"inventory_report_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return send_file(pdf_io, mimetype='application/pdf',
                     as_attachment=True, download_name=filename)


@products_bp.route('/export/excel')
@login_required
def export_excel():
    if not _require_manager():
        return redirect(url_for('products.index'))

    search = request.args.get('search', '')
    category_id = request.args.get('category_id', type=int)

    query = Product.query
    if search:
        query = query.filter(or_(
            Product.name.contains(search),
            Product.sku.contains(search),
            Product.barcode.contains(search)
        ))
    if category_id:
        query = query.filter(Product.category_id == category_id)

    products = query.order_by(Product.category_id, Product.name).all()

    # --- Company info ---
    company_name    = SystemSetting.get('company_name', 'Electronics Store POS')
    company_address = SystemSetting.get('company_address', '')
    company_phone   = SystemSetting.get('company_phone', '')

    # --- Stats ---
    total_products  = len(products)
    total_units     = sum(p.quantity_in_stock for p in products)
    low_stock_count = sum(1 for p in products if p.is_low_stock)
    wholesale_value = sum((p.cost_price or Decimal('0')) * p.quantity_in_stock for p in products)
    retail_value    = sum((p.selling_price or Decimal('0')) * p.quantity_in_stock for p in products)
    potential_margin = retail_value - wholesale_value

    # --- Category breakdown ---
    cat_map = defaultdict(lambda: {'products': 0, 'units': 0, 'wholesale': Decimal('0'), 'retail': Decimal('0')})
    for p in products:
        cat_name = p.category.name if p.category else 'Uncategorized'
        cat_map[cat_name]['products'] += 1
        cat_map[cat_name]['units']    += p.quantity_in_stock
        cat_map[cat_name]['wholesale'] += (p.cost_price or Decimal('0')) * p.quantity_in_stock
        cat_map[cat_name]['retail']    += (p.selling_price or Decimal('0')) * p.quantity_in_stock
    category_breakdown = sorted(cat_map.items(), key=lambda x: x[1]['retail'], reverse=True)

    # ── Style helpers ──
    NAVY     = "1A3A5C"
    WHITE    = "FFFFFF"
    LIGHT_BG = "F7FAFC"
    GREEN_BG = "C6F6D5"
    GREEN_FG = "22543D"
    RED_BG   = "FED7D7"
    RED_FG   = "742A2A"
    YELLOW_BG= "FEFCBF"
    YELLOW_FG= "744210"
    BLUE_BG  = "EBF8FF"
    BLUE_FG  = "1A3A5C"
    BORDER_COLOR = "E2E8F0"

    def thin_border():
        s = Side(style='thin', color=BORDER_COLOR)
        return Border(left=s, right=s, top=s, bottom=s)

    def make_header_cell(ws_ref, row, col, value, bg=NAVY, fg=WHITE):
        cell = ws_ref.cell(row=row, column=col, value=value)
        cell.font      = Font(bold=True, color=fg, size=10)
        cell.fill      = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border()
        return cell

    def style_data_cell(cell, align="left", bold=False, color=None, bg=None, num_fmt=None):
        cell.font      = Font(bold=bold, color=(color or "2D3748"), size=10)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border    = thin_border()
        if bg:
            cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        if num_fmt:
            cell.number_format = num_fmt
        return cell

    def autofit_columns(ws_ref, min_w=8, max_w=40):
        for col_cells in ws_ref.columns:
            best = min_w
            for c in col_cells:
                try:
                    v = str(c.value) if c.value is not None else ''
                    best = max(best, min(len(v) + 2, max_w))
                except Exception:
                    pass
            ws_ref.column_dimensions[get_column_letter(col_cells[0].column)].width = best

    wb = openpyxl.Workbook()

    # ══════════════════════════════════════════════
    # SHEET 1: Summary
    # ══════════════════════════════════════════════
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.sheet_view.showGridLines = False

    # Company header
    ws_sum.merge_cells("A1:D1")
    c = ws_sum.cell(row=1, column=1, value=company_name.upper())
    c.font      = Font(bold=True, color=NAVY, size=14)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 28

    ws_sum.merge_cells("A2:D2")
    c2 = ws_sum.cell(row=2, column=1, value=f"Inventory Report — {datetime.now().strftime('%B %d, %Y')}")
    c2.font      = Font(color="718096", size=11)
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[2].height = 20

    if company_address or company_phone:
        meta = " | ".join(x for x in [company_address, company_phone] if x)
        ws_sum.merge_cells("A3:D3")
        c3 = ws_sum.cell(row=3, column=1, value=meta)
        c3.font      = Font(color="A0AEC0", size=9)
        c3.alignment = Alignment(horizontal="center", vertical="center")
        ws_sum.row_dimensions[3].height = 16

    # Summary metrics
    start_row = 5
    make_header_cell(ws_sum, start_row, 1, "Metric")
    make_header_cell(ws_sum, start_row, 2, "Value")
    ws_sum.column_dimensions['A'].width = 34
    ws_sum.column_dimensions['B'].width = 22
    ws_sum.row_dimensions[start_row].height = 18

    rows_data = [
        ("Total Products",              total_products,   None,    None),
        ("Total Units",                 total_units,      None,    None),
        ("Inventory Value (Wholesale)", float(wholesale_value), None, '"$"#,##0.00'),
        ("Inventory Value (Retail)",    float(retail_value),    BLUE_BG, '"$"#,##0.00'),
        ("Potential Margin",            float(potential_margin),GREEN_BG, '"$"#,##0.00'),
        ("Products OK",                 total_products - low_stock_count, None, None),
        ("Products Low Stock",          low_stock_count, (RED_BG if low_stock_count > 0 else None), None),
    ]
    for i, (label, value, row_bg, num_fmt) in enumerate(rows_data, start=start_row + 1):
        ws_sum.row_dimensions[i].height = 16
        lc = ws_sum.cell(row=i, column=1, value=label)
        vc = ws_sum.cell(row=i, column=2, value=value)
        alt = LIGHT_BG if i % 2 == 0 else None
        style_data_cell(lc, bold=True, bg=row_bg or alt)
        style_data_cell(vc, align="right", bg=row_bg or alt, num_fmt=num_fmt)

    # ══════════════════════════════════════════════
    # SHEET 2: Category Breakdown
    # ══════════════════════════════════════════════
    ws_cat = wb.create_sheet("Category Breakdown")
    ws_cat.sheet_view.showGridLines = False

    cat_headers = ["Category", "Products", "Units", "Wholesale Value", "Retail Value", "Margin", "Margin %"]
    for col_i, h in enumerate(cat_headers, 1):
        make_header_cell(ws_cat, 1, col_i, h)
    ws_cat.row_dimensions[1].height = 18
    ws_cat.column_dimensions['A'].width = 24
    ws_cat.column_dimensions['B'].width = 12
    ws_cat.column_dimensions['C'].width = 12
    ws_cat.column_dimensions['D'].width = 20
    ws_cat.column_dimensions['E'].width = 20
    ws_cat.column_dimensions['F'].width = 18
    ws_cat.column_dimensions['G'].width = 12

    for i, (cat_name, data) in enumerate(category_breakdown, start=2):
        ws_cat.row_dimensions[i].height = 15
        alt = LIGHT_BG if i % 2 == 0 else None
        margin = data['retail'] - data['wholesale']
        margin_pct = (float(margin) / float(data['wholesale']) * 100) if data['wholesale'] > 0 else 0

        cells_vals = [
            (cat_name, "left",  True,  alt, None),
            (data['products'], "center", False, alt, None),
            (data['units'],    "center", False, alt, None),
            (float(data['wholesale']), "right", False, alt, '"$"#,##0.00'),
            (float(data['retail']),    "right", False, alt, '"$"#,##0.00'),
            (float(margin),            "right", False, GREEN_BG if margin > 0 else alt, '"$"#,##0.00'),
            (round(margin_pct, 1),     "right", False, GREEN_BG if margin_pct >= 15 else (YELLOW_BG if margin_pct > 0 else alt), '0.0"%"'),
        ]
        for col_i, (val, align, bold, bg, num_fmt) in enumerate(cells_vals, 1):
            c = ws_cat.cell(row=i, column=col_i, value=val)
            style_data_cell(c, align=align, bold=bold, bg=bg, num_fmt=num_fmt)

    # ══════════════════════════════════════════════
    # SHEET 3: Complete Inventory
    # ══════════════════════════════════════════════
    ws_inv = wb.create_sheet("Inventory")
    ws_inv.sheet_view.showGridLines = False

    inv_headers = [
        "ID", "Product Name", "Category", "SKU", "Barcode",
        "Cost Price", "Selling Price", "Wholesale Price",
        "Profit/Unit", "Margin %", "Stock", "Low Stock Threshold", "Status"
    ]
    for col_i, h in enumerate(inv_headers, 1):
        make_header_cell(ws_inv, 1, col_i, h)
    ws_inv.row_dimensions[1].height = 18

    col_widths = [8, 32, 18, 14, 16, 14, 14, 16, 14, 11, 8, 18, 12]
    for col_i, w in enumerate(col_widths, 1):
        ws_inv.column_dimensions[get_column_letter(col_i)].width = w

    for i, p in enumerate(products, start=2):
        ws_inv.row_dimensions[i].height = 15
        alt  = LIGHT_BG if i % 2 == 0 else None
        cost  = float(p.cost_price or 0)
        price = float(p.selling_price or 0)
        whl   = float(p.wholesale_price or 0) if p.wholesale_price else None
        profit = price - cost
        margin_pct = round(profit / cost * 100, 1) if cost > 0 else 0
        status = "Low Stock" if p.is_low_stock else "In Stock"

        margin_bg = GREEN_BG if margin_pct >= 30 else (YELLOW_BG if margin_pct >= 10 else (RED_BG if cost > 0 else alt))
        status_bg = RED_BG if p.is_low_stock else GREEN_BG
        status_fg = RED_FG if p.is_low_stock else GREEN_FG

        row_data = [
            (p.id,                           "center", False, alt,      None,            None),
            (p.name,                         "left",   True,  alt,      None,            None),
            (p.category.name if p.category else '-', "left", False, alt, None,          None),
            (p.sku,                          "left",   False, alt,      None,            None),
            (p.barcode or '-',               "left",   False, alt,      None,            None),
            (cost,                           "right",  False, alt,      '"$"#,##0.00',   None),
            (price,                          "right",  False, alt,      '"$"#,##0.00',   None),
            (whl,                            "right",  False, alt,      '"$"#,##0.00',   None),
            (profit,                         "right",  False, GREEN_BG if profit > 0 else alt, '"$"#,##0.00', None),
            (margin_pct,                     "right",  True,  margin_bg, '0.0"%"',       None),
            (p.quantity_in_stock,            "center", False, alt,      None,            None),
            (p.low_stock_threshold,          "center", False, alt,      None,            None),
            (status,                         "center", True,  status_bg, None,           status_fg),
        ]
        for col_i, (val, align, bold, bg, num_fmt, fg) in enumerate(row_data, 1):
            c = ws_inv.cell(row=i, column=col_i, value=val)
            style_data_cell(c, align=align, bold=bold, bg=bg, num_fmt=num_fmt,
                            color=(fg or "2D3748"))

    # ══════════════════════════════════════════════
    # SHEET 4: Low Stock
    # ══════════════════════════════════════════════
    low_products = [p for p in products if p.is_low_stock]
    if low_products:
        ws_low = wb.create_sheet("Low Stock Alerts")
        ws_low.sheet_view.showGridLines = False

        low_headers = ["ID", "Product Name", "Category", "SKU", "Stock", "Threshold", "Cost Price", "Selling Price"]
        for col_i, h in enumerate(low_headers, 1):
            make_header_cell(ws_low, 1, col_i, h, bg="C53030", fg=WHITE)
        ws_low.row_dimensions[1].height = 18
        low_widths = [8, 32, 18, 14, 8, 10, 14, 14]
        for col_i, w in enumerate(low_widths, 1):
            ws_low.column_dimensions[get_column_letter(col_i)].width = w

        for i, p in enumerate(low_products, start=2):
            ws_low.row_dimensions[i].height = 15
            row_data = [
                (p.id,                                   "center"),
                (p.name,                                 "left"),
                (p.category.name if p.category else '-', "left"),
                (p.sku,                                  "left"),
                (p.quantity_in_stock,                    "center"),
                (p.low_stock_threshold,                  "center"),
                (float(p.cost_price or 0),               "right"),
                (float(p.selling_price or 0),            "right"),
            ]
            for col_i, (val, align) in enumerate(row_data, 1):
                c = ws_low.cell(row=i, column=col_i, value=val)
                num_fmt = '"$"#,##0.00' if col_i >= 7 else None
                style_data_cell(c, align=align, bg=RED_BG, color=RED_FG, num_fmt=num_fmt)

    excel_io = BytesIO()
    wb.save(excel_io)
    excel_io.seek(0)

    filename = f"inventory_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        excel_io,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ===========================================================================
# Image helpers
# ===========================================================================

def _save_product_image(file_data):
    original_ext = file_data.filename.rsplit('.', 1)[-1].lower()
    unique_filename = f"{uuid.uuid4().hex[:12]}.{original_ext}"
    upload_folder = current_app.config['PRODUCT_IMAGES_UPLOAD_FOLDER']
    file_data.save(os.path.join(upload_folder, unique_filename))
    return unique_filename


def _delete_product_image(filename):
    try:
        filepath = os.path.join(current_app.config['PRODUCT_IMAGES_UPLOAD_FOLDER'], filename)
        if os.path.exists(filepath):
            os.remove(filepath)
    except Exception as e:
        current_app.logger.warning(f"Failed to delete product image {filename}: {e}")
